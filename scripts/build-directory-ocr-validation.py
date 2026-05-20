"""
Re-parse the original image PDF with Windows OCR and compare it with:

1. the existing parsed workbook
2. current production members

Outputs are local review artifacts only. The script does not update Supabase.
"""
from __future__ import annotations

import csv
import io
import json
import os
import re
import subprocess
import sys
import tempfile
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

import fitz
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

REPO_ROOT = Path(__file__).resolve().parents[1]
ENV_PATH = REPO_ROOT / "chflow-app" / ".env.local"
OUT_ROOT = REPO_ROOT / "MS_AX" / "generated" / f"directory_ocr_validation_{date.today().isoformat()}"
IMAGE_DIR = OUT_ROOT / "page_images"
OCR_DIR = OUT_ROOT / "ocr_json"
CROP_DIR = OUT_ROOT / "difference_crops"

ORIGINAL_PDF_SIZE = 36_979_781
EXISTING_WORKBOOK_SIZE = 144_544
FIRST_DIRECTORY_PAGE = 44
LAST_DIRECTORY_PAGE = 111
RENDER_SCALE = 2

USER_CONFIRMED_DB_KEEP_IDS = {
    # source_page is the pasture/household listing page. A spouse may be
    # missing from that page or placed in another pasture; MDB spouse data and
    # role/photo-page OCR can still confirm the relationship.
    "0a620439-7644-4383-b45c-9d3d3bb0d5e2",  # 노성희: spouse 강희열 confirmed
    "6cab1b79-e934-4e37-a830-db4c2a6253f8",  # 남정희: 김성환 spouse, no photo confirmed
    "fae1ccb5-e56e-48d9-bc65-e3fe77f4ee26",  # 김성환: 남정희 spouse, photo only confirmed
    "f49301b2-afbe-4085-a552-e5ba8ccfe520",  # 임태섭: spouse 전민자, no photo confirmed
    "f3318f4e-e1ba-4cfd-a4af-2179e7a37d1f",  # 박기완: 김현숙 spouse, no photo confirmed
    "f91e1ac8-fc5e-4af2-9bf9-383a35f8ea97",  # 김현숙: 시무권사, spouse 박기완 confirmed
    "34398216-107a-4759-a016-b50edfab769b",  # 박영석: 김현숙/박기완 child confirmed
    "0245c1d8-96fd-474e-a343-1931bbced9bf",  # 권관옥: phone/spouse confirmed
    "201c686b-11cc-4f97-a8a1-cb19e1b2b006",  # 이석철: 은퇴시무집사 confirmed
    "901e2d9a-e925-4d32-8ad3-857a1bda3e40",  # 임순이: phone/spouse confirmed
    "d04dc94e-932a-48a0-8ee8-7ce1efdb5af7",  # 김찬규: was misread as 전진규, spouse 김은지 confirmed
    "c5ef4139-499b-4d89-b60b-59482ab70e1c",  # 전진규: spouse 구영교 confirmed
    "43a9f9b4-3d0c-440a-bb1f-fd32dc4a0827",  # 구영교: spouse 전진규 confirmed
    "68e6467f-c329-4519-9cdf-c8941bbdb5c4",  # 박영란: 은퇴시무권사, spouse 신승원 confirmed
    "5a582395-1d2b-4b85-8e71-e238279f6ef4",  # 신승원: 은퇴시무집사, spouse 박영란 confirmed
    "86f3b692-04c4-4731-a74d-dcc3b19fc79a",  # 조유태: 은퇴시무집사 confirmed
    "33b5418d-cb93-4ca5-8f71-fd165b65bf41",  # 박현지: 서리집사, spouse 진영우 confirmed
    "fc07afc3-6497-4804-9cd8-0f49e266f4db",  # 진영우: 서리집사, spouse 박현지 confirmed
    "fbe0acfb-0252-4cd6-a292-8c999f0448dc",  # 박순자: p88/MDB family 1050 corrected from 박소자
    "f16b71c9-beeb-451c-a4b8-b91a63dc99f1",  # 안용기: spouse 박순자 confirmed
    "e4d5f7dc-baab-470f-a087-6e87608e0c9b",  # 안민호: 안용기/박순자 child confirmed
    "a8b00caa-6c31-4203-ac9d-ba22c9173772",  # 안준호: 안용기/박순자 child confirmed
    "3a31276a-b0de-41d7-8613-64ce0acb4586",  # 김인기: OCR/manual crop corrected from 강민기
    "e747c5c6-8474-40dc-a4e3-a737b2acf315",  # 박영석: OCR/manual crop corrected from 박영식
    "caf89ec6-8233-4795-b1c5-698bf55eba91",  # 임순현: OCR/manual crop corrected from 임순혜
    "b0e146c3-f48a-4234-a17e-a04848523c1a",  # 안준: 집사 confirmed
    "04971639-0fcb-4cab-aab4-54baf57e7c32",  # 황영애: 권사 confirmed
    "9518d3e1-1db6-4377-ad6a-f267718ac7d4",  # 한현용: spouse/shared contact OCR false conflict
    "38dd1bc8-d0be-4d4f-b824-2c097ee55736",  # 이일호: 성도, spouse 조향선 confirmed
    "51561ecf-deff-42fd-a6ba-0927747b90d5",  # 조향선: 서리집사, spouse 이일호 confirmed
    "1129547b-d1a2-4784-8d8e-584fc0a7b5d1",  # 윤경숙: p92 original corrected from 공준숙
    "6f433f08-d26d-443b-a899-06f6c1d7cc45",  # 박순형: p92 original corrected from 박승형
    "38864201-b30b-4a79-921f-57bf3fe50793",  # 박은정: no spouse, children 진성/재성, 김지원 unrelated
    "009f6bdb-65d8-4603-8e00-fbe93c43694e",  # 김지원: unrelated to 박은정, shared phone/OCR conflict resolved
    "50588029-f69e-419e-ad1a-a62e91ebac3d",  # 진성: 박은정 child confirmed
    "b7056234-965d-4371-83c7-a2bdcd046d15",  # 재성: 박은정 child confirmed
    "b4b0f9f4-2e4f-40d6-ab23-bfc8a172cb5d",  # 한영진: active member, OCR candidate ignored
    "adcec55a-a524-444a-be0f-6a11a9151c74",  # 신영이: 시무권사, spouse 김지호 confirmed
    "ddd4ef61-b067-44f7-863a-1eb7bc7a96ab",  # 김지호: 성도, spouse 신영이 confirmed
    "d60bf316-7040-463a-ad2d-e74cf21a80ba",  # 천미숙: 시무권사, spouse 이정필 confirmed
    "9b5fd4e2-a5da-4a6e-b697-affef2588f34",  # 이정필: 성도, spouse 천미숙 confirmed
    "d0d108fd-a19d-42d1-846f-2a0d7a0c60bc",  # 이용순: 교육사 confirmed
    "f3a802c9-3c82-4e75-a890-5fcb51eae435",  # 김성현: 정영교 spouse, no photo confirmed
    "664bd15d-ca60-424f-baa0-75c534c6eeb0",  # 신관익: original p78 corrected from OCR 신관옥
    "9285c644-a516-4c1f-8798-0c965e14d591",  # 백정희: original p78 corrected from OCR 박정희
    "238379a6-40ae-4faf-8994-c54c2f7edba4",  # 신은정: 신관익/백정희 child confirmed
    "f526a8ee-8397-4cdd-91fc-1e27c6560e52",  # 김은순: p94 original corrected from 장수화
    "2e26ba2f-b2a7-4a00-9d42-652e03de7dea",  # 전송수: spouse 김은순 confirmed
    "47ad29d7-bfb1-4283-a97f-8e93716f3616",  # 박수진: p97 original corrected from 박옥진
    "e3647779-c5f5-45c5-9c4a-6dcaf31a3287",  # 선우석: spouse 박수진 confirmed
    "dee865f3-e050-4266-be41-40a9c0753f37",  # 정사무엘: p97 original corrected from 박해준
    "84744504-238d-419f-b289-593699b1e6b7",  # 손미혜: p97 original corrected from 정주원
    "33d53645-fac4-496c-b132-6071e481767c",  # 장준호: p97 original corrected from 정순호
    "6de31251-f09e-4e0c-9f60-3f1a373fa85b",  # 김수진: spouse 장준호 confirmed
    "74e03b92-35b0-492d-8c77-b8d8fde07353",  # 정희은: p97 original corrected from 정화원
    "d0ac186c-1754-47c6-b5d8-8432b33e91d0",  # 노상용: OCR/manual crop confirms DB row
    "96b8734a-7c62-4a85-92a2-b90d1185d25c",  # 문인옥: OCR/manual crop confirms DB row
    "c55d1bcd-846e-4b03-8725-ad09936c4acf",  # 쯔엉티프엉: OCR/manual crop confirms DB row
    "366dd0cd-3991-4d8a-b32b-1deb87cddb55",  # 연무진: OCR/manual crop confirms DB row
    "3b2da086-73cd-47a4-bbe3-120002e236d2",  # 김건희: OCR/manual crop corrected from duplicate 김근수
    "46914331-75c0-421a-b7c6-cecef31fe09e",  # 이영신: OCR/manual crop corrected from 이명선
    "87941a31-9c10-4ffa-af08-a255c12fef3c",  # 박선옥: OCR/manual crop corrected from 주성철
    "ceb92bfb-4fac-4d53-9005-5a4917e7775b",  # 박정숙: OCR/manual crop corrected from 송혜숙
    "d9902676-e879-4673-ba52-7986fb6a299d",  # 박강민: OCR/manual crop corrected from 박강인
    "e14d6409-421c-4fff-a96e-6d10fb16b8fe",  # 박명학: p99 original corrected from 박성혁
    "5b78a57f-07c4-4360-94b5-27de7222f4b5",  # 조은혜: spouse 박명학 confirmed
    "7d7b0492-2307-4058-b86a-a42a26b82c81",  # 김상석: p99 original corrected from 박재혁
    "5d21c7ef-6446-41cd-bb08-db883d62c364",  # 정수연: spouse 김상석 confirmed
    "e3fa8121-1f42-4f49-9209-fd65854f9fe4",  # 정외섭: p101 original corrected from 정희성
    "b0d49282-46b5-4206-beb4-5eb5f7b7cbfa",  # 박화선: spouse 정외섭 confirmed
    "073431d1-ebc8-44c4-80f3-e01e0e05acca",  # 조경자: spouse 김규호 confirmed
    "91b8a869-c356-44bf-94b9-8e780295c73d",  # 박상범: p110 no phone, OCR phone belonged to 이예소
    "7f9fad94-8488-4d12-8d76-2f9e7dcae5c1",  # 이예소: p110 phone confirmed
    "cab7e2da-ccef-4763-8ad2-71ef11f0eb39",  # 정성령: p110 original corrected from 장성영
    "f1733b7d-d61d-42b5-a939-af9575ff7999",  # 김홍남: 은퇴시무집사 confirmed
    "ac5fbb3a-de60-47c2-82dd-7573802547d5",  # 김근수: 은퇴시무집사 confirmed
    "0db5745d-e948-41d7-abd2-fb62f3837104",  # 조석호: 교육사 confirmed
}

ROLE_FIXES = {
    "Al": "시",
    "kl": "시",
    "A": "사",
    "ㅅ": "사",
    "口": "무",
    "_": "-",
}

KNOWN_ROLES = [
    "은퇴시무집사",
    "명예시무집사",
    "시무집사",
    "서리집사",
    "명예집사",
    "은퇴장로",
    "원로장로",
    "시무장로",
    "명예장로",
    "은퇴권사",
    "명예권사",
    "시무권사",
    "장로",
    "권사",
    "집사",
    "청년",
    "목사",
    "사모",
    "전도사",
    "교육사",
]

ROLE_EQUIV = {
    "은회시무집사": "은퇴시무집사",
    "은퇴시무집사": "은퇴시무집사",
    "시무집사": "시무집사",
    "서리집사": "서리집사",
    "명예집사": "명예집사",
    "은퇴장로": "은퇴장로",
    "원로장로": "원로장로",
    "시무장로": "시무장로",
    "은퇴권사": "은퇴권사",
    "명예권사": "명예권사",
    "시무권사": "시무권사",
    "장로": "장로",
    "권사": "권사",
    "집사": "집사",
    "청년": "청년",
}


@dataclass
class ParsedRow:
    source: str
    page: int
    row_no: int | None
    y: float | None
    name: str
    sub_role: str
    phone: str
    home_phone: str
    plain_name: str = ""
    grassland_name: str = ""
    pasture_name: str = ""
    raw_role: str = ""
    raw_phone: str = ""
    note: str = ""


def load_env() -> None:
    if not ENV_PATH.exists():
        return
    for raw in ENV_PATH.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def rest_headers() -> dict[str, str]:
    key = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
    return {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
    }


def encode_rest_path(path: str) -> str:
    if "?" not in path:
        return path
    base, query = path.split("?", 1)
    return base + "?" + urllib.parse.quote(query, safe="=&.,%*!():")


def fetch_all(table: str, select: str = "*", extra: str = "") -> list[dict[str, Any]]:
    url = os.environ["NEXT_PUBLIC_SUPABASE_URL"]
    rows: list[dict[str, Any]] = []
    page = 0
    page_size = 1000
    while True:
        path = f"/{table}?select={select}"
        if extra:
            path += "&" + extra
        req = urllib.request.Request(
            f"{url}/rest/v1{encode_rest_path(path)}",
            headers={
                **rest_headers(),
                "Range-Unit": "items",
                "Range": f"{page * page_size}-{(page + 1) * page_size - 1}",
            },
        )
        try:
            with urllib.request.urlopen(req, timeout=120) as response:
                chunk = json.loads(response.read())
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")[:500]
            raise SystemExit(f"Supabase REST error {exc.code} for {table}: {detail}") from exc
        if not chunk:
            break
        rows.extend(chunk)
        if len(chunk) < page_size:
            break
        page += 1
    return rows


def find_file_by_size(root: Path, suffix: str, size: int, prefer_parts: tuple[str, ...] = ()) -> Path:
    matches = [path for path in root.rglob(f"*{suffix}") if path.stat().st_size == size]
    if not matches:
        raise SystemExit(f"Could not find {suffix} with size {size}")
    return sorted(
        matches,
        key=lambda path: (
            0 if any(part in path.parts for part in prefer_parts) else 1,
            len(path.parts),
            str(path),
        ),
    )[0]


def norm_name(value: Any) -> str:
    text = "" if value is None else str(value)
    hangul = re.findall(r"[가-힣]+", text)
    return "".join(hangul)


def norm_phone(value: Any) -> str:
    return re.sub(r"\D", "", "" if value is None else str(value))


def clean_phone(value: Any) -> str:
    digits = norm_phone(value)
    if len(digits) == 11:
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    if len(digits) == 10 and digits.startswith("02"):
        return f"{digits[:2]}-{digits[2:6]}-{digits[6:]}"
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    return digits


def normalize_role(value: Any) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"\s+", "", text)
    for bad, good in ROLE_FIXES.items():
        text = text.replace(bad, good)
    if re.search(r"은.?퇴.*시.?무.*집.?사", text):
        return "은퇴시무집사"
    if re.search(r"은.?퇴.*치.?무.*집.?사", text):
        return "은퇴시무집사"
    if re.search(r"은.?퇴.*시.?집", text):
        return "은퇴시무집사"
    if re.search(r"시.?무.*집.?사", text):
        return "시무집사"
    if re.search(r"명.?예.*집.?사", text):
        return "명예집사"
    if re.search(r"은.?퇴.*권.?사", text):
        return "은퇴권사"
    if re.search(r"명.?예.*권.?사", text):
        return "명예권사"
    if re.search(r"은.?퇴.*장.?로", text):
        return "은퇴장로"
    if re.search(r"원.?로.*장.?로", text):
        return "원로장로"
    for bad, good in ROLE_EQUIV.items():
        if bad in text:
            return good
    for role in KNOWN_ROLES:
        if role in text:
            return role
    return ""


def render_original_pdf(pdf_path: Path) -> None:
    IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    doc = fitz.open(pdf_path)
    for page_no in range(FIRST_DIRECTORY_PAGE, LAST_DIRECTORY_PAGE + 1):
        out = IMAGE_DIR / f"p{page_no:03d}.png"
        if out.exists():
            continue
        pix = doc[page_no - 1].get_pixmap(matrix=fitz.Matrix(RENDER_SCALE, RENDER_SCALE), alpha=False)
        pix.save(out)


def run_windows_ocr() -> None:
    OCR_DIR.mkdir(parents=True, exist_ok=True)
    script = r"""
param([string]$ImageDir,[string]$OutDir)
$ErrorActionPreference='Stop'
Add-Type -AssemblyName System.Runtime.WindowsRuntime
[Windows.Storage.StorageFile, Windows.Storage, ContentType=WindowsRuntime] | Out-Null
[Windows.Storage.FileAccessMode, Windows.Storage, ContentType=WindowsRuntime] | Out-Null
[Windows.Graphics.Imaging.BitmapDecoder, Windows.Graphics.Imaging, ContentType=WindowsRuntime] | Out-Null
[Windows.Media.Ocr.OcrEngine, Windows.Foundation, ContentType=WindowsRuntime] | Out-Null
[Windows.Globalization.Language, Windows.Globalization, ContentType=WindowsRuntime] | Out-Null
function Await-Operation($op, [Type]$resultType) {
  $method = [System.WindowsRuntimeSystemExtensions].GetMethods() | Where-Object {
    $_.Name -eq 'AsTask' -and $_.IsGenericMethod -and $_.GetParameters().Count -eq 1
  } | Select-Object -First 1
  $task = $method.MakeGenericMethod($resultType).Invoke($null, @($op))
  $task.Wait() | Out-Null
  return $task.Result
}
New-Item -ItemType Directory -Force -Path $OutDir | Out-Null
$engine = [Windows.Media.Ocr.OcrEngine]::TryCreateFromLanguage([Windows.Globalization.Language]::new('ko'))
if ($null -eq $engine) { throw 'Korean Windows OCR engine is unavailable.' }
foreach($img in Get-ChildItem -Path $ImageDir -Filter *.png | Sort-Object Name){
  $outPath = Join-Path $OutDir ($img.BaseName + '.json')
  if (Test-Path $outPath) { Write-Output "skip $($img.Name)"; continue }
  $file = Await-Operation ([Windows.Storage.StorageFile]::GetFileFromPathAsync($img.FullName)) ([Windows.Storage.StorageFile])
  $stream = Await-Operation ($file.OpenAsync([Windows.Storage.FileAccessMode]::Read)) ([Windows.Storage.Streams.IRandomAccessStream])
  $decoder = Await-Operation ([Windows.Graphics.Imaging.BitmapDecoder]::CreateAsync($stream)) ([Windows.Graphics.Imaging.BitmapDecoder])
  $bitmap = Await-Operation ($decoder.GetSoftwareBitmapAsync()) ([Windows.Graphics.Imaging.SoftwareBitmap])
  $result = Await-Operation ($engine.RecognizeAsync($bitmap)) ([Windows.Media.Ocr.OcrResult])
  $lines = New-Object System.Collections.Generic.List[object]
  foreach ($line in $result.Lines) {
    $words = New-Object System.Collections.Generic.List[object]
    foreach ($w in $line.Words) {
      $r = $w.BoundingRect
      $words.Add([pscustomobject]@{ text=$w.Text; x=[double]$r.X; y=[double]$r.Y; width=[double]$r.Width; height=[double]$r.Height })
    }
    $lines.Add([pscustomobject]@{ text=$line.Text; words=$words })
  }
  $json=$lines | ConvertTo-Json -Depth 8
  [IO.File]::WriteAllText($outPath,$json,[Text.UTF8Encoding]::new($false))
  Write-Output "ocr $($img.Name)"
}
"""
    with tempfile.NamedTemporaryFile("w", suffix=".ps1", delete=False, encoding="utf-8") as handle:
        handle.write(script)
        ps_path = handle.name
    try:
        subprocess.run(
            [
                "powershell",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                ps_path,
                "-ImageDir",
                str(IMAGE_DIR),
                "-OutDir",
                str(OCR_DIR),
            ],
            cwd=str(REPO_ROOT),
            check=True,
        )
    finally:
        Path(ps_path).unlink(missing_ok=True)


def line_xy(line: dict[str, Any]) -> tuple[float, float]:
    words = line.get("words") or []
    if not words:
        return 0.0, 0.0
    return min(float(word["x"]) for word in words), min(float(word["y"]) for word in words)


def parse_page_ocr(path: Path) -> list[ParsedRow]:
    page = int(path.stem[1:])
    lines = json.loads(path.read_text(encoding="utf-8"))
    enriched = []
    for line in lines:
        x, y = line_xy(line)
        text = str(line.get("text") or "").strip()
        if text:
            enriched.append({"text": text, "x": x, "y": y})

    header_y_values = [
        line["y"]
        for line in enriched
        if "성명" in line["text"] and ("번호" in line["text"] or 360 <= line["x"] <= 430)
    ]
    if not header_y_values:
        return []
    header_y = min(header_y_values)

    title_lines = [line for line in enriched if line["y"] < header_y - 35 and line["x"] < 500]
    title_texts = [line["text"] for line in sorted(title_lines, key=lambda item: (item["y"], item["x"]))]
    plain = next((text for text in title_texts if "평원" in text), "")
    pasture = next((text for text in title_texts if "목장" in text), "")
    grassland = next((text for text in title_texts if text not in {plain, pasture} and 2 <= len(text) <= 20), "")

    names = []
    for line in enriched:
        if line["y"] <= header_y + 20:
            continue
        if not (360 <= line["x"] <= 430):
            continue
        name = norm_name(line["text"])
        if 2 <= len(name) <= 5 and name not in {"성명", "번호", "가족"}:
            names.append({**line, "name": name})

    role_lines = [line for line in enriched if line["y"] > header_y + 20 and 430 <= line["x"] <= 520]
    phone_lines = [
        line
        for line in enriched
        if line["y"] > header_y + 20
        and line["x"] >= 720
        and re.search(r"\d{2,3}[-_ .:]?\d{3,4}[-_ .:]?\d{4}", line["text"])
    ]
    number_lines = [
        line
        for line in enriched
        if line["y"] > header_y + 20 and 320 <= line["x"] <= 360 and re.fullmatch(r"\d{1,2}", line["text"])
    ]

    def nearest(items: list[dict[str, Any]], y: float, tolerance: float = 16) -> dict[str, Any] | None:
        candidates = [(abs(item["y"] - y), item) for item in items if abs(item["y"] - y) <= tolerance]
        if not candidates:
            return None
        return sorted(candidates, key=lambda pair: pair[0])[0][1]

    rows: list[ParsedRow] = []
    seen: set[tuple[str, int, int]] = set()
    for name_line in sorted(names, key=lambda item: (item["y"], item["x"])):
        role_line = nearest(role_lines, name_line["y"], 18)
        phone_line = nearest(phone_lines, name_line["y"], 18)
        number_line = nearest(number_lines, name_line["y"], 18)
        role_raw = role_line["text"] if role_line else ""
        phone_raw = phone_line["text"] if phone_line else ""
        row_key = (name_line["name"], page, round(float(name_line["y"])))
        if row_key in seen:
            continue
        seen.add(row_key)
        rows.append(
            ParsedRow(
                source="windows_ocr_original_pdf",
                page=page,
                row_no=int(number_line["text"]) if number_line else None,
                y=float(name_line["y"]),
                name=name_line["name"],
                sub_role=normalize_role(role_raw),
                phone=clean_phone(phone_raw),
                home_phone="",
                plain_name=plain,
                grassland_name=grassland,
                pasture_name=pasture,
                raw_role=role_raw,
                raw_phone=phone_raw,
                note="" if role_raw else "role_missing",
            )
        )
    return rows


def load_existing_rows(path: Path) -> list[ParsedRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows: list[ParsedRow] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        page = row[18]
        name = norm_name(row[6])
        if not page or not name:
            continue
        rows.append(
            ParsedRow(
                source="existing_parsed_workbook",
                page=int(page),
                row_no=int(row[0]) if row[0] else None,
                y=None,
                name=name,
                sub_role=normalize_role(row[9]),
                phone=clean_phone(row[15]),
                home_phone=clean_phone(row[14]),
                plain_name=str(row[1] or ""),
                grassland_name=str(row[3] or ""),
                pasture_name=str(row[4] or ""),
                raw_role=str(row[9] or ""),
                raw_phone=str(row[15] or ""),
                note=str(row[19] or ""),
            )
        )
    return rows


def write_parsed_workbook(rows: list[ParsedRow], path: Path) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "새 OCR 파싱"
    headers = [
        "source",
        "page",
        "row_no",
        "y",
        "name",
        "sub_role",
        "phone",
        "home_phone",
        "plain",
        "grassland",
        "pasture",
        "raw_role",
        "raw_phone",
        "note",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E0F2FE")
    for row in rows:
        ws.append([
            row.source,
            row.page,
            row.row_no,
            row.y,
            row.name,
            row.sub_role,
            row.phone,
            row.home_phone,
            row.plain_name,
            row.grassland_name,
            row.pasture_name,
            row.raw_role,
            row.raw_phone,
            row.note,
        ])
    ws.freeze_panes = "A2"
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(32, max(10, max(len(str(cell.value or "")) for cell in col) + 2))
    workbook.save(path)


def index_rows(rows: list[ParsedRow]) -> tuple[dict[tuple[int, str], list[ParsedRow]], dict[tuple[int, str], list[ParsedRow]], dict[int, list[ParsedRow]]]:
    by_phone: dict[tuple[int, str], list[ParsedRow]] = defaultdict(list)
    by_name: dict[tuple[int, str], list[ParsedRow]] = defaultdict(list)
    by_page: dict[int, list[ParsedRow]] = defaultdict(list)
    for row in rows:
        for phone in {norm_phone(row.phone), norm_phone(row.home_phone)} - {""}:
            by_phone[(row.page, phone)].append(row)
        by_name[(row.page, norm_name(row.name))].append(row)
        by_page[row.page].append(row)
    return by_phone, by_name, by_page


def find_match(member: dict[str, Any], by_phone: dict[tuple[int, str], list[ParsedRow]], by_name: dict[tuple[int, str], list[ParsedRow]], by_page: dict[int, list[ParsedRow]]) -> tuple[ParsedRow | None, str]:
    page = int(member.get("source_page") or 0)
    name = norm_name(member.get("name"))
    phones = []
    for phone in [norm_phone(member.get("phone")), norm_phone(member.get("home_phone"))]:
        if phone and phone not in phones:
            phones.append(phone)
    for phone in phones:
        candidates = by_phone.get((page, phone), [])
        if candidates:
            candidates = sorted(candidates, key=lambda row: (name_distance(name, row.name), row.y or 0))
            return candidates[0], "page_phone"
    candidates = by_name.get((page, norm_name(member.get("name"))), [])
    if candidates:
        return candidates[0], "page_name"
    candidates = [row for row in by_page.get(page, []) if name_close(name, row.name)]
    if candidates:
        candidates = sorted(candidates, key=lambda row: (name_distance(name, row.name), row.y or 0))
        return candidates[0], "page_name_fuzzy"
    return None, "no_match"


def role_equal(left: Any, right: Any) -> bool:
    left_role = normalize_role(left)
    right_role = normalize_role(right)
    if left_role == right_role:
        return True
    equivalent_pairs = {
        frozenset(("집사", "서리집사")),
        frozenset(("권사", "시무권사")),
    }
    return frozenset((left_role, right_role)) in equivalent_pairs


def name_equal(left: Any, right: Any) -> bool:
    return norm_name(left) == norm_name(right)


def name_distance(left: Any, right: Any) -> int:
    a = norm_name(left)
    b = norm_name(right)
    if not a:
        return len(b)
    if not b:
        return len(a)
    previous = list(range(len(b) + 1))
    for i, a_char in enumerate(a, start=1):
        current = [i]
        for j, b_char in enumerate(b, start=1):
            current.append(min(
                previous[j] + 1,
                current[j - 1] + 1,
                previous[j - 1] + (0 if a_char == b_char else 1),
            ))
        previous = current
    return previous[-1]


def name_close(left: Any, right: Any) -> bool:
    a = norm_name(left)
    b = norm_name(right)
    return bool(a and b and abs(len(a) - len(b)) <= 1 and name_distance(a, b) <= 1)


def identity_supported_with_unreadable_role(member: dict[str, Any], row: ParsedRow | None) -> bool:
    return bool(row and not row.sub_role and (name_equal(member.get("name"), row.name) or name_close(member.get("name"), row.name)))


def classify(member: dict[str, Any], old: ParsedRow | None, new: ParsedRow | None) -> str:
    if str(member.get("id")) in USER_CONFIRMED_DB_KEEP_IDS:
        return "user_confirmed_db"

    db_name = member.get("name")
    db_role = member.get("sub_role")
    old_name_ok = old is not None and name_equal(db_name, old.name)
    new_name_ok = new is not None and name_equal(db_name, new.name)
    old_name_close = old is not None and name_close(db_name, old.name)
    new_name_close = new is not None and name_close(db_name, new.name)
    old_role_ok = old is not None and role_equal(db_role, old.sub_role)
    new_role_ok = new is not None and role_equal(db_role, new.sub_role)

    if old is None and new is None:
        return "no_pdf_match"
    if old_name_ok and new_name_ok and old_role_ok and new_role_ok:
        return "all_agree"
    if old and new and name_equal(old.name, new.name) and name_close(db_name, new.name) and len(norm_name(new.name)) == len(norm_name(db_name)) and not name_equal(db_name, new.name):
        return "name_review_ab_agree_db_diff"
    if old and new and old_name_ok and new_name_ok and old.sub_role and new.sub_role and role_equal(old.sub_role, new.sub_role) and not role_equal(db_role, new.sub_role):
        return "role_review_ab_agree_db_diff"
    if new_name_ok and new_role_ok:
        return "new_ocr_matches_db"
    if old_name_ok and old_role_ok:
        return "existing_parse_matches_db"
    if new_name_close and not new_name_ok and new_role_ok:
        return "new_ocr_fuzzy_name_matches_db"
    if old_name_close and not old_name_ok and old_role_ok:
        return "existing_parse_fuzzy_name_matches_db"
    if identity_supported_with_unreadable_role(member, new) or identity_supported_with_unreadable_role(member, old):
        return "identity_supported_role_unreadable"
    if (new_name_ok and new and new.sub_role) or (old_name_ok and old and old.sub_role):
        return "single_parse_role_review"
    if (new_name_close and new and new.sub_role) or (old_name_close and old and old.sub_role):
        return "single_parse_name_role_review"
    return "ambiguous"


def family_shared_contact_conflict(
    member: dict[str, Any],
    old: ParsedRow | None,
    new: ParsedRow | None,
    household_names: set[str],
) -> bool:
    candidate_names = [
        row.name
        for row in (old, new)
        if row and row.name and not name_close(member.get("name"), row.name)
    ]
    if not candidate_names:
        return False

    spouse_name = norm_name(member.get("spouse_name"))
    family_names = {norm_name(name) for name in household_names if name}
    if spouse_name:
        family_names.add(spouse_name)
    family_names.discard(norm_name(member.get("name")))

    return bool(family_names) and all(norm_name(name) in family_names for name in candidate_names)


def crop_filename(row: dict[str, Any]) -> str:
    if row["verdict"] == "all_agree":
        return ""
    page_no = int(row["source_page"] or 0)
    if not (FIRST_DIRECTORY_PAGE <= page_no <= LAST_DIRECTORY_PAGE):
        return ""
    if row.get("new_y") == "":
        return ""
    safe_name = norm_name(row["db_name"]) or "member"
    return f"p{page_no:03d}_{safe_name}_{row['member_id']}.png"


def build_comparison(existing_rows: list[ParsedRow], new_rows: list[ParsedRow]) -> list[dict[str, Any]]:
    load_env()
    members = fetch_all(
        "members",
        select="id,name,status,phone,home_phone,family_church,sub_role,spouse_name,is_child,source_page,household_id",
        extra="status=eq.active&source_page=not.is.null&order=source_page.asc,name.asc",
    )
    existing_by_phone, existing_by_name, existing_by_page = index_rows(existing_rows)
    new_by_phone, new_by_name, new_by_page = index_rows(new_rows)
    names_by_household: dict[str, set[str]] = defaultdict(set)
    for member in members:
        household_id = str(member.get("household_id") or "")
        if household_id and member.get("name"):
            names_by_household[household_id].add(str(member.get("name")))

    output = []
    for member in members:
        if member.get("is_child"):
            continue
        old, old_method = find_match(member, existing_by_phone, existing_by_name, existing_by_page)
        new, new_method = find_match(member, new_by_phone, new_by_name, new_by_page)
        verdict = classify(member, old, new)
        if verdict == "ambiguous" and family_shared_contact_conflict(
            member,
            old,
            new,
            names_by_household.get(str(member.get("household_id") or ""), set()),
        ):
            verdict = "family_shared_contact"
        review_bucket, review_reason = describe_verdict(verdict)
        comparison_row = {
            "verdict": verdict,
            "review_bucket": review_bucket,
            "review_reason": review_reason,
            "member_id": member.get("id"),
            "source_page": member.get("source_page"),
            "db_name": member.get("name"),
            "db_sub_role": normalize_role(member.get("sub_role")),
            "db_spouse_name": member.get("spouse_name"),
            "db_phone": member.get("phone"),
            "db_home_phone": member.get("home_phone"),
            "existing_match_method": old_method,
            "existing_name": old.name if old else "",
            "existing_sub_role": old.sub_role if old else "",
            "existing_phone": old.phone if old else "",
            "existing_raw_role": old.raw_role if old else "",
            "new_match_method": new_method,
            "new_name": new.name if new else "",
            "new_sub_role": new.sub_role if new else "",
            "new_phone": new.phone if new else "",
            "new_raw_role": new.raw_role if new else "",
            "new_y": new.y if new else "",
            "new_note": new.note if new else "",
        }
        comparison_row["crop_file"] = crop_filename(comparison_row)
        output.append(comparison_row)
    return output


def describe_verdict(verdict: str) -> tuple[str, str]:
    descriptions = {
        "all_agree": ("auto_pass", "existing parse, new OCR, and DB agree"),
        "user_confirmed_db": ("auto_pass", "DB value was explicitly confirmed despite OCR disagreement"),
        "new_ocr_matches_db": ("auto_pass", "new OCR exactly supports current DB"),
        "existing_parse_matches_db": ("auto_pass", "existing parse exactly supports current DB"),
        "new_ocr_fuzzy_name_matches_db": ("auto_pass", "new OCR has a one-character name variation but role supports current DB"),
        "existing_parse_fuzzy_name_matches_db": ("auto_pass", "existing parse has a one-character name variation but role supports current DB"),
        "identity_supported_role_unreadable": ("low_priority", "name/phone identity is supported, but role cell was unreadable or blank"),
        "name_review_ab_agree_db_diff": ("visual_review", "both parses agree on a different close name"),
        "role_review_ab_agree_db_diff": ("visual_review", "both parses agree on a different role"),
        "single_parse_role_review": ("visual_review", "one parse supports the identity, but role differs from DB"),
        "single_parse_name_role_review": ("visual_review", "one parse has a close name candidate and role differs from DB"),
        "family_shared_contact": ("low_priority", "OCR matched a spouse/household member through shared contact info"),
        "no_pdf_match": ("manual_review", "no reliable row matched in either parse"),
        "ambiguous": ("manual_review", "matched evidence conflicts or points to another household member"),
    }
    return descriptions.get(verdict, ("manual_review", "unclassified disagreement"))


def write_comparison_workbook(rows: list[dict[str, Any]], path: Path) -> Path:
    auto_pass = {"all_agree", "user_confirmed_db", "new_ocr_matches_db", "existing_parse_matches_db", "new_ocr_fuzzy_name_matches_db", "existing_parse_fuzzy_name_matches_db"}
    visual_review = {"name_review_ab_agree_db_diff", "role_review_ab_agree_db_diff", "single_parse_role_review", "single_parse_name_role_review"}
    low_priority = {"identity_supported_role_unreadable", "family_shared_contact"}
    manual_review = {"ambiguous", "no_pdf_match"}
    user_review_headers = [
        "user_decision",
        "confirmed_name",
        "confirmed_sub_role",
        "photo_note",
        "user_note",
    ]

    workbook = Workbook()
    ws = workbook.active
    ws.title = "3자 비교"
    headers = list(rows[0].keys()) if rows else []

    def write_sheet(sheet: Any, sheet_rows: list[dict[str, Any]], fill: str, include_user_review: bool = False) -> None:
        sheet_headers = headers + (user_review_headers if include_user_review else [])
        sheet.append(sheet_headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor=fill)
        for row in sheet_rows:
            values = [row.get(header, "") for header in headers]
            if include_user_review:
                values.extend(["", "", "", "", ""])
            sheet.append(values)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    write_sheet(ws, rows, "FEF3C7")
    diff = workbook.create_sheet("차이 후보")
    write_sheet(diff, [row for row in rows if row["verdict"] != "all_agree"], "FEE2E2")
    visual = workbook.create_sheet("시각 확인 후보")
    write_sheet(visual, [row for row in rows if row["verdict"] in visual_review], "DBEAFE", include_user_review=True)
    manual = workbook.create_sheet("수동 검수 요청")
    write_sheet(manual, [row for row in rows if row["verdict"] in manual_review], "EDE9FE", include_user_review=True)
    low = workbook.create_sheet("낮은 우선순위")
    write_sheet(low, [row for row in rows if row["verdict"] in low_priority], "E0E7FF", include_user_review=True)
    passed = workbook.create_sheet("자동 통과")
    write_sheet(passed, [row for row in rows if row["verdict"] in auto_pass], "DCFCE7")

    for sheet in workbook.worksheets:
        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = min(34, max(10, max(len(str(cell.value or "")) for cell in col) + 2))
    try:
        workbook.save(path)
        return path
    except PermissionError:
        fallback = path.with_name(f"{path.stem}_review_columns{path.suffix}")
        workbook.save(fallback)
        return fallback


def write_diff_crops(pdf_path: Path, comparison_rows: list[dict[str, Any]]) -> None:
    CROP_DIR.mkdir(parents=True, exist_ok=True)
    for old_crop in CROP_DIR.glob("*.png"):
        old_crop.unlink()
    doc = fitz.open(pdf_path)
    for row in comparison_rows:
        if row["verdict"] == "all_agree":
            continue
        page_no = int(row["source_page"] or 0)
        if not (FIRST_DIRECTORY_PAGE <= page_no <= LAST_DIRECTORY_PAGE):
            continue
        y = row.get("new_y")
        if y == "":
            continue
        y_pdf = float(y) / RENDER_SCALE
        page = doc[page_no - 1]
        rect = fitz.Rect(150, max(0, y_pdf - 22), 470, min(page.rect.height, y_pdf + 34))
        pix = page.get_pixmap(matrix=fitz.Matrix(2.5, 2.5), clip=rect, alpha=False)
        filename = row.get("crop_file") or crop_filename(row)
        if not filename:
            continue
        out = CROP_DIR / filename
        pix.save(out)


def write_summary(rows: list[dict[str, Any]], parsed_count: int, path: Path) -> None:
    counts = Counter(row["verdict"] for row in rows)
    bucket_counts = Counter(row["review_bucket"] for row in rows)
    lines = [
        f"# Directory OCR Validation - {date.today().isoformat()}",
        "",
        "## Inputs",
        "",
        "- Original image PDF: `자료/데이터베이스 추출용.pdf`",
        "- Existing parsed workbook: `명성교회_회원DB_검수용_v2.1.xlsx`",
        "- Production DB: active members with `source_page`",
        "- OCR engine: Windows Korean OCR, local machine",
        "",
        "## Outputs",
        "",
        f"- New OCR parsed rows: {parsed_count}",
        f"- Compared DB adult rows: {len(rows)}",
        f"- Difference crop directory: `{CROP_DIR}`",
        "",
        "## Verdict Counts",
        "",
    ]
    for key, count in counts.most_common():
        lines.append(f"- `{key}`: {count}")
    lines.extend([
        "",
        "## Review Bucket Counts",
        "",
    ])
    for key, count in bucket_counts.most_common():
        lines.append(f"- `{key}`: {count}")
    lines.extend([
        "",
        "## Workbook Sheets",
        "",
        "- `3자 비교`: every compared DB row.",
        "- `차이 후보`: every row except full agreement.",
        "- `시각 확인 후보`: strict name/role candidates that should be checked against crop images before any DB update.",
        "- `수동 검수 요청`: ambiguous rows and rows with no reliable PDF match.",
        "- `낮은 우선순위`: identity is supported, but role evidence is unreadable or blank.",
        "- `자동 통과`: rows where the current DB is supported by at least one parse path.",
        "",
        "## User Review Columns",
        "",
        "- `user_decision`: `DB유지`, `요람반영`, `직접수정`, or `보류`.",
        "- `confirmed_name`: final name when it should be changed or explicitly confirmed.",
        "- `confirmed_sub_role`: final role when it should be changed or explicitly confirmed.",
        "- `photo_note`: photo-specific note such as `사진없음` or `배우자 하옥련 사진 75p`.",
        "- `user_note`: any other context needed before DB update.",
        "",
        "## Interpretation",
        "",
        "- `all_agree`: existing parse, new OCR parse, and DB agree on name/role.",
        "- `user_confirmed_db`: DB value was explicitly confirmed despite OCR disagreement.",
        "- `new_ocr_matches_db`: new OCR agrees with DB while existing parse differs or is incomplete.",
        "- `existing_parse_matches_db`: existing parse agrees with DB while new OCR differs or is incomplete.",
        "- `new_ocr_fuzzy_name_matches_db`: new OCR has a one-character name variation, but role supports the DB.",
        "- `existing_parse_fuzzy_name_matches_db`: existing parse has a one-character name variation, but role supports the DB.",
        "- `identity_supported_role_unreadable`: name/phone identity is supported, but role evidence is unreadable or blank.",
        "- `name_review_ab_agree_db_diff`: both parses agree on a different name; requires visual confirmation before any DB change.",
        "- `role_review_ab_agree_db_diff`: both parses agree on a different role; good candidate for visual confirmation.",
        "- `single_parse_role_review`: one parse supports the identity, but role differs from DB.",
        "- `single_parse_name_role_review`: one parse has a close name candidate and role differs from DB.",
        "- `ambiguous`: disagreement pattern is not mechanically safe.",
    ])
    path.write_text("\n".join(lines), encoding="utf-8-sig")


def main() -> None:
    OUT_ROOT.mkdir(parents=True, exist_ok=True)
    pdf_path = find_file_by_size(REPO_ROOT, ".pdf", ORIGINAL_PDF_SIZE, prefer_parts=("자료",))
    existing_path = find_file_by_size(REPO_ROOT / "MS_AX", ".xlsx", EXISTING_WORKBOOK_SIZE)

    print(f"original_pdf={pdf_path}")
    print(f"existing_workbook={existing_path}")
    render_original_pdf(pdf_path)
    run_windows_ocr()

    new_rows: list[ParsedRow] = []
    for path in sorted(OCR_DIR.glob("p*.json")):
        new_rows.extend(parse_page_ocr(path))
    new_rows.sort(key=lambda row: (row.page, row.y or 0, row.name))

    existing_rows = load_existing_rows(existing_path)
    comparison_rows = build_comparison(existing_rows, new_rows)
    comparison_rows.sort(key=lambda row: (row["source_page"] or 0, row["db_name"] or ""))

    parsed_xlsx = OUT_ROOT / f"directory_new_windows_ocr_parse_{date.today().isoformat()}.xlsx"
    comparison_xlsx = OUT_ROOT / f"directory_ocr_3way_comparison_{date.today().isoformat()}.xlsx"
    summary_md = OUT_ROOT / f"directory_ocr_validation_summary_{date.today().isoformat()}.md"
    write_parsed_workbook(new_rows, parsed_xlsx)
    comparison_xlsx = write_comparison_workbook(comparison_rows, comparison_xlsx)
    write_diff_crops(pdf_path, comparison_rows)
    write_summary(comparison_rows, len(new_rows), summary_md)

    print(f"new_rows={len(new_rows)}")
    print("verdict_counts=" + json.dumps(Counter(row["verdict"] for row in comparison_rows), ensure_ascii=False))
    print(f"parsed_xlsx={parsed_xlsx}")
    print(f"comparison_xlsx={comparison_xlsx}")
    print(f"summary={summary_md}")
    print(f"crops={CROP_DIR}")


if __name__ == "__main__":
    main()

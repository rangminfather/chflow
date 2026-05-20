"""
Build a read-only sub_role normalization report.

The report compares current production members with the parsed directory PDF
workbook and emits only conservative candidates. It does not update Supabase.
"""
from __future__ import annotations

import csv
import io
import json
import os
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

REPO_ROOT = Path(__file__).resolve().parents[1]
ENV_PATH = REPO_ROOT / "chflow-app" / ".env.local"
PARSED_WORKBOOK_SIZE = 144544
OUT_DIR = REPO_ROOT / "MS_AX" / "generated"

SAFE_NORMALIZATIONS = {
    ("집사", "서리집사"): "generic_jipsa_to_seori_jipsa",
    ("집사/시무집사", "시무집사"): "jipsa_simoo_jipsa_to_simoo_jipsa",
    ("권사", "시무권사"): "generic_kwonsa_to_simoo_kwonsa",
    ("장로", "시무장로"): "generic_jangro_to_simoo_jangro",
}

EXPLICIT_ROLES = {
    "은퇴권사",
    "명예권사",
    "은퇴장로",
    "원로장로",
    "명예집사",
}


@dataclass(frozen=True)
class ParsedRow:
    row_no: int
    name: str
    spouse_name: str
    family_church: str
    sub_role: str
    mobile_phone: str
    home_phone: str
    source_page: int | None
    plain_name: str
    grassland_name: str
    pasture_name: str
    note: str


def load_env() -> None:
    if not ENV_PATH.exists():
        raise SystemExit(f"Missing env file: {ENV_PATH}")

    for raw in ENV_PATH.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def rest_headers() -> dict[str, str]:
    key = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
    return {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
    }


def encode_rest_path(path: str) -> str:
    if "?" not in path:
        return path
    base, query = path.split("?", 1)
    return base + "?" + urllib.parse.quote(query, safe="=&.,%*!():")


def fetch_all(table: str, select: str = "*", extra: str = "") -> list[dict[str, Any]]:
    url = os.environ["NEXT_PUBLIC_SUPABASE_URL"]
    rows: list[dict[str, Any]] = []
    page = 0
    page_size = 1000

    while True:
        path = f"/{table}?select={select}"
        if extra:
            path += "&" + extra
        req = urllib.request.Request(
            f"{url}/rest/v1{encode_rest_path(path)}",
            headers={
                **rest_headers(),
                "Range-Unit": "items",
                "Range": f"{page * page_size}-{(page + 1) * page_size - 1}",
            },
        )
        try:
            with urllib.request.urlopen(req, timeout=120) as response:
                chunk = json.loads(response.read())
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")[:500]
            raise SystemExit(f"Supabase REST error {exc.code} for {table}: {detail}") from exc

        if not chunk:
            break
        rows.extend(chunk)
        if len(chunk) < page_size:
            break
        page += 1

    return rows


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def norm_name(value: Any) -> str:
    return re.sub(r"\s+", "", text(value))


def norm_phone(value: Any) -> str:
    return re.sub(r"\D", "", text(value))


def norm_role(value: Any) -> str:
    value = text(value)
    value = re.sub(r"\s+", "", value)
    return value


def parse_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def find_parsed_workbook() -> Path:
    candidates = [
        path
        for path in (REPO_ROOT / "MS_AX").rglob("*.xlsx")
        if path.stat().st_size == PARSED_WORKBOOK_SIZE
    ]
    if not candidates:
        raise SystemExit("Could not find parsed directory workbook.")
    return candidates[0]


def load_parsed_rows(path: Path) -> list[ParsedRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["회원 마스터"]
    headers = [text(cell) for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    indexes = {name: idx for idx, name in enumerate(headers)}

    required = ["#", "평원", "초원", "목장", "이름", "배우자", "가정교회", "직분", "자택전화", "휴대폰", "페이지", "비고"]
    missing = [name for name in required if name not in indexes]
    if missing:
        raise SystemExit(f"Parsed workbook missing columns: {', '.join(missing)}")

    rows: list[ParsedRow] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        name = text(row[indexes["이름"]])
        if not name:
            continue
        rows.append(
            ParsedRow(
                row_no=parse_int(row[indexes["#"]]) or 0,
                name=name,
                spouse_name=text(row[indexes["배우자"]]),
                family_church=text(row[indexes["가정교회"]]),
                sub_role=norm_role(row[indexes["직분"]]),
                mobile_phone=text(row[indexes["휴대폰"]]),
                home_phone=text(row[indexes["자택전화"]]),
                source_page=parse_int(row[indexes["페이지"]]),
                plain_name=text(row[indexes["평원"]]),
                grassland_name=text(row[indexes["초원"]]),
                pasture_name=text(row[indexes["목장"]]),
                note=text(row[indexes["비고"]]),
            )
        )
    return rows


def member_phone_values(member: dict[str, Any]) -> set[str]:
    values = {
        norm_phone(member.get("phone")),
        norm_phone(member.get("home_phone")),
    }
    return {value for value in values if value}


def parsed_phone_values(row: ParsedRow) -> set[str]:
    values = {
        norm_phone(row.mobile_phone),
        norm_phone(row.home_phone),
    }
    return {value for value in values if value}


def build_matches(members: list[dict[str, Any]], parsed_rows: list[ParsedRow]) -> list[dict[str, Any]]:
    rows_by_name_page: dict[tuple[str, int | None], list[ParsedRow]] = defaultdict(list)
    for row in parsed_rows:
        rows_by_name_page[(norm_name(row.name), row.source_page)].append(row)

    matches: list[dict[str, Any]] = []
    for member in members:
        if member.get("source_page") is None:
            continue
        if member.get("is_child"):
            continue

        key = (norm_name(member.get("name")), parse_int(member.get("source_page")))
        candidates = rows_by_name_page.get(key, [])
        member_phones = member_phone_values(member)
        scored: list[tuple[int, ParsedRow, str]] = []
        for row in candidates:
            row_phones = parsed_phone_values(row)
            phone_overlap = sorted(member_phones & row_phones)
            if phone_overlap:
                scored.append((100, row, "name_page_phone"))
            elif len(candidates) == 1:
                scored.append((80, row, "unique_name_page"))

        if not scored:
            continue
        scored.sort(key=lambda item: item[0], reverse=True)
        score, parsed, confidence = scored[0]
        if len(scored) > 1 and scored[1][0] == score:
            confidence = "ambiguous_name_page"

        current_role = norm_role(member.get("sub_role"))
        parsed_role = parsed.sub_role
        if not parsed_role or current_role == parsed_role:
            continue

        reason = SAFE_NORMALIZATIONS.get((current_role, parsed_role))
        category = "manual_review"
        if reason and confidence in {"name_page_phone", "unique_name_page"}:
            category = "safe_auto"
        elif current_role in EXPLICIT_ROLES:
            reason = "explicit_current_role_keep_unless_directly_verified"
        elif parsed_role in EXPLICIT_ROLES:
            reason = "explicit_pdf_role_requires_manual_review"
        else:
            reason = "role_family_or_value_change_requires_manual_review"

        matches.append(
            {
                "category": category,
                "reason": reason,
                "confidence": confidence,
                "member_id": member.get("id"),
                "name": member.get("name"),
                "status": member.get("status"),
                "source_page": member.get("source_page"),
                "current_sub_role": current_role,
                "pdf_sub_role": parsed_role,
                "family_church": member.get("family_church"),
                "pdf_family_church": parsed.family_church,
                "phone": member.get("phone"),
                "home_phone": member.get("home_phone"),
                "pdf_mobile_phone": parsed.mobile_phone,
                "pdf_home_phone": parsed.home_phone,
                "spouse_name": member.get("spouse_name"),
                "pdf_spouse_name": parsed.spouse_name,
                "plain_name": parsed.plain_name,
                "grassland_name": parsed.grassland_name,
                "pasture_name": parsed.pasture_name,
                "pdf_row_no": parsed.row_no,
                "pdf_note": parsed.note,
            }
        )

    return matches


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    fieldnames = [
        "category",
        "reason",
        "confidence",
        "member_id",
        "name",
        "status",
        "source_page",
        "current_sub_role",
        "pdf_sub_role",
        "family_church",
        "pdf_family_church",
        "phone",
        "home_phone",
        "pdf_mobile_phone",
        "pdf_home_phone",
        "spouse_name",
        "pdf_spouse_name",
        "plain_name",
        "grassland_name",
        "pasture_name",
        "pdf_row_no",
        "pdf_note",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_markdown(path: Path, rows: list[dict[str, Any]], workbook: Path, member_count: int, parsed_count: int) -> None:
    category_counts = Counter(row["category"] for row in rows)
    reason_counts = Counter(row["reason"] for row in rows)
    safe_rows = [row for row in rows if row["category"] == "safe_auto"]

    lines = [
        f"# Sub-role Normalization Report - {date.today().isoformat()}",
        "",
        "## Scope",
        "",
        "- Source DB: production Supabase via service-role REST, read-only.",
        f"- Parsed PDF workbook: `{workbook}`",
        f"- Production members fetched: {member_count}",
        f"- Parsed workbook rows loaded: {parsed_count}",
        "- Matching rule: same normalized name and same directory page; phone overlap preferred, unique name/page allowed.",
        "- No database writes were performed.",
        "",
        "## Summary",
        "",
        f"- Total differing direct matches: {len(rows)}",
        f"- Safe automatic candidates: {category_counts.get('safe_auto', 0)}",
        f"- Manual review candidates: {category_counts.get('manual_review', 0)}",
        "",
        "## Reasons",
        "",
    ]
    if reason_counts:
        for reason, count in reason_counts.most_common():
            lines.append(f"- `{reason}`: {count}")
    else:
        lines.append("- No differing candidates found.")

    lines.extend(["", "## Safe Automatic Candidates", ""])
    if safe_rows:
        lines.append("| name | page | current | target | confidence | reason |")
        lines.append("| --- | ---: | --- | --- | --- | --- |")
        for row in safe_rows[:100]:
            lines.append(
                f"| {row['name']} | {row['source_page']} | {row['current_sub_role']} | "
                f"{row['pdf_sub_role']} | {row['confidence']} | `{row['reason']}` |"
            )
        if len(safe_rows) > 100:
            lines.append(f"| ... | ... | ... | ... | ... | {len(safe_rows) - 100} more in CSV |")
    else:
        lines.append("- None.")

    lines.extend(
        [
            "",
            "## Next Step",
            "",
            "Review the CSV detail before applying any updates. If the safe_auto set is accepted, generate a narrow SQL update only for those member IDs.",
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8-sig")


def main() -> None:
    load_env()
    workbook = find_parsed_workbook()
    parsed_rows = load_parsed_rows(workbook)
    members = fetch_all(
        "members",
        select="id,name,status,phone,home_phone,family_church,sub_role,spouse_name,is_child,source_page",
        extra="status=eq.active&order=source_page.asc,name.asc",
    )
    rows = build_matches(members, parsed_rows)
    rows.sort(key=lambda row: (row["category"] != "safe_auto", row["source_page"] or 9999, row["name"] or ""))

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = date.today().isoformat()
    csv_path = OUT_DIR / f"sub_role_normalization_candidates_{stamp}.csv"
    md_path = OUT_DIR / f"sub_role_normalization_report_{stamp}.md"
    write_csv(csv_path, rows)
    write_markdown(md_path, rows, workbook, len(members), len(parsed_rows))

    counts = Counter(row["category"] for row in rows)
    print(f"members={len(members)} parsed_rows={len(parsed_rows)} differing_matches={len(rows)}")
    print(f"safe_auto={counts.get('safe_auto', 0)} manual_review={counts.get('manual_review', 0)}")
    print(f"csv={csv_path}")
    print(f"report={md_path}")


if __name__ == "__main__":
    main()

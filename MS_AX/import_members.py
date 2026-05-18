"""
성도정보 업로드 (xlsx → Supabase)

사용:
  python import_members.py <엑셀파일>
  python import_members.py <엑셀파일> --mode 3
  python import_members.py <엑셀파일> --mode 2 --yes

모드:
  3 (⭐⭐ 가장 추천) — 차이 미리보기 후 확인 → 적용 (UPDATE + INSERT, 삭제 X)
  2 (⭐ 추천)        — id 일치 행 UPDATE만 (신규/삭제 무시) — 가장 안전
  1 (⚠️  비추천)     — 전체 덮어쓰기 (엑셀에 없는 행 = DB 삭제) — 사고 위험

대상 시트:
  Members      — id 매칭 (필수)
  Relations    — id 매칭, kind/role 갱신
  Ministries   — id 매칭
  Directory(households) — id 매칭, 주소/order_no 갱신만 지원

* 엑셀 _name 컬럼 (plain_name/pasture_name 등) 은 무시됨 — household_id 가 진실.
"""
import sys, io, os, json, urllib.request, urllib.error, urllib.parse
import argparse
from datetime import datetime, date

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

try:
    from openpyxl import load_workbook
except ImportError:
    print('[ERROR] openpyxl 필요: pip install openpyxl')
    sys.exit(1)


# ============================================================
# 환경
# ============================================================
ENV_PATH = 'c:/csh/project/chflow/chflow-app/.env.local'
with open(ENV_PATH, 'r', encoding='utf-8') as f:
    for line in f:
        line = line.strip()
        if '=' in line and not line.startswith('#'):
            k, v = line.split('=', 1)
            os.environ[k] = v

URL = os.environ['NEXT_PUBLIC_SUPABASE_URL']
KEY = os.environ['SUPABASE_SERVICE_ROLE_KEY']
H = {
    'apikey': KEY,
    'Authorization': f'Bearer {KEY}',
    'Content-Type': 'application/json',
    'Prefer': 'return=minimal',
}


def _enc(p):
    if '?' in p:
        b, q = p.split('?', 1)
        return b + '?' + urllib.parse.quote(q, safe='=&.,%*!')
    return p


def http(method, path, body=None, headers=None):
    url = f'{URL}/rest/v1{_enc(path)}'
    data = json.dumps(body, ensure_ascii=False, default=str).encode('utf-8') if body is not None else None
    h = {**H, **(headers or {})}
    req = urllib.request.Request(url, data=data, headers=h, method=method)
    try:
        with urllib.request.urlopen(req, timeout=120) as r:
            raw = r.read()
            return json.loads(raw) if raw else None
    except urllib.error.HTTPError as e:
        msg = e.read().decode('utf-8', errors='replace')
        print(f'  [HTTP {e.code}] {method} {path}: {msg[:300]}')
        return None


def fetch_all(table, select='*'):
    rows, page, SIZE = [], 0, 1000
    while True:
        url = f'{URL}/rest/v1/{table}?select={select}'
        req = urllib.request.Request(url, headers={
            **H,
            'Range-Unit': 'items',
            'Range': f'{page * SIZE}-{(page + 1) * SIZE - 1}',
        })
        try:
            with urllib.request.urlopen(req, timeout=120) as r:
                chunk = json.loads(r.read())
        except urllib.error.HTTPError as e:
            print(f'[ERR] fetch {table}: {e}')
            return []
        if not chunk:
            break
        rows.extend(chunk)
        if len(chunk) < SIZE:
            break
        page += 1
    return rows


# ============================================================
# 시트 → 행 변환
# ============================================================
SHEET_SCHEMAS = {
    'Members': {
        'table': 'members',
        # 엑셀 컬럼 → DB 컬럼 (None이면 무시)
        'cols': {
            'id': 'id',
            'excel_row_no': 'excel_row_no',
            'name': 'name',
            'gender': 'gender',
            'birth_date': 'birth_date',
            'phone': 'phone',
            'family_church': 'family_church',
            'sub_role': 'sub_role',
            'spouse_name': 'spouse_name',
            'plain_name': None,        # 표시용 — 무시
            'grassland_name': None,
            'pasture_name': None,
            'address': None,           # households 에 있음
            'is_child': 'is_child',
            'guard_status': 'guard_status',
            'has_account': None,       # 파생값
            'photo_status': 'photo_status',
            'photo_page': 'photo_page',
            'photo_url': 'photo_url',
            'source_page': 'source_page',
            'notes': 'notes',
            'household_id': 'household_id',
            'spouse_id': 'spouse_id',
        },
        'pk': 'id',
    },
    'Relations': {
        'table': 'member_relations',
        'cols': {
            'id': 'id',
            'subject_id': 'subject_id',
            'subject_name': None,
            'relative_id': 'relative_id',
            'relative_name': None,
            'kind': 'kind',
            'role': 'role',
        },
        'pk': 'id',
    },
    'Ministries': {
        'table': 'member_ministries',
        'cols': {
            'id': 'id',
            'member_id': 'member_id',
            'member_name': None,
            'ministry': 'ministry',
            'role': 'role',
            'notes': 'notes',
        },
        'pk': 'id',
    },
    'Directory': {
        'table': 'households',
        'cols': {
            'household_id': 'id',
            'plain_name': None,
            'grassland_name': None,
            'pasture_name': None,
            'address': 'address',
            'home_phone': 'home_phone',
            'order_no': 'order_no',
            'pasture_id': 'pasture_id',
            'grassland_id': None,
            'plain_id': None,
        },
        'pk': 'id',
    },
}


def _norm(v):
    """엑셀 → DB 정규화. 빈 문자열 → None, datetime → ISO date."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return v


def read_sheet(ws, schema):
    headers = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in r):
            continue
        rec = {}
        for col_name, db_col in schema['cols'].items():
            if db_col is None:
                continue
            try:
                idx = headers.index(col_name)
            except ValueError:
                continue
            rec[db_col] = _norm(r[idx])
        rows.append(rec)
    return rows


# ============================================================
# Diff
# ============================================================
def compute_diff(excel_rows, db_rows, pk='id'):
    """엑셀 vs DB 비교. 반환: updates, inserts, deletes, unchanged."""
    db_map = {r[pk]: r for r in db_rows}
    excel_ids = set()
    updates, inserts, unchanged = [], [], []
    for er in excel_rows:
        eid = er.get(pk)
        if not eid:
            inserts.append(er)
            continue
        excel_ids.add(eid)
        dr = db_map.get(eid)
        if dr is None:
            inserts.append(er)
            continue
        # 비교: 엑셀에 있는 컬럼만
        changed = {}
        for k, v in er.items():
            if k == pk:
                continue
            db_v = dr.get(k)
            # date 정규화
            if isinstance(db_v, str) and 'T' in db_v:
                db_v = db_v.split('T')[0]
            if (v or '') != (db_v or '') and not (v is None and db_v is None):
                changed[k] = (db_v, v)
        if changed:
            updates.append({'id': eid, 'name': dr.get('name') or dr.get('subject_id') or '?',
                            'changes': changed, 'row': er})
        else:
            unchanged.append(er)

    deletes = [r for pk_, r in db_map.items() if pk_ not in excel_ids]
    return updates, inserts, deletes, unchanged


# ============================================================
# 적용
# ============================================================
def apply_updates(table, updates):
    ok = 0
    for u in updates:
        body = {k: v for k, v in u['row'].items() if k != 'id'}
        res = http('PATCH', f'/{table}?id=eq.{u["id"]}', body)
        if res is not None:
            ok += 1
        else:
            print(f'  [실패] UPDATE {u["id"]} ({u.get("name")})')
    return ok


def apply_inserts(table, inserts):
    if not inserts:
        return 0
    ok = 0
    for i in range(0, len(inserts), 100):
        chunk = [{k: v for k, v in r.items() if k != 'id' or v} for r in inserts[i:i + 100]]
        res = http('POST', f'/{table}', chunk)
        if res is not None:
            ok += len(chunk)
    return ok


def apply_deletes(table, deletes, pk='id'):
    ok = 0
    for r in deletes:
        rid = r.get(pk)
        if not rid:
            continue
        res = http('DELETE', f'/{table}?{pk}=eq.{rid}')
        if res is not None:
            ok += 1
    return ok


# ============================================================
# 모드별 실행
# ============================================================
def show_diff_summary(sheet, updates, inserts, deletes, unchanged, mode):
    print(f'\n--- [{sheet}] 변경 요약 ---')
    print(f'  변경(UPDATE): {len(updates)}')
    print(f'  추가(INSERT): {len(inserts)}'
          + ('' if mode != 2 else '   ← 모드2에서 무시'))
    print(f'  삭제(DELETE): {len(deletes)}'
          + ('' if mode == 1 else '   ← 모드' + str(mode) + '에서 무시'))
    print(f'  변화없음:    {len(unchanged)}')

    if updates:
        print('\n  변경 예시 (최대 10건):')
        for u in updates[:10]:
            chs = ', '.join(f'{k}: {a!r}→{b!r}' for k, (a, b) in u['changes'].items())
            print(f'    · {u.get("name", u["id"][:8])}  {chs}')
        if len(updates) > 10:
            print(f'    ... 외 {len(updates) - 10}건')

    if inserts and mode != 2:
        print('\n  추가 예시 (최대 5건):')
        for r in inserts[:5]:
            print(f'    · {r.get("name") or r}')

    if deletes and mode == 1:
        print('\n  ⚠️  삭제 예시 (최대 5건):')
        for r in deletes[:5]:
            print(f'    · {r.get("name") or r.get("id")}')


def run_sheet(sheet, schema, excel_rows, mode, auto_yes):
    table = schema['table']
    pk = schema['pk']
    print(f'\n=== [{sheet}] {table} 처리 ===')
    print(f'  엑셀 행: {len(excel_rows)}')

    db_rows = fetch_all(table)
    print(f'  DB 행:   {len(db_rows)}')

    updates, inserts, deletes, unchanged = compute_diff(excel_rows, db_rows, pk)
    show_diff_summary(sheet, updates, inserts, deletes, unchanged, mode)

    if mode == 2:
        inserts, deletes = [], []
    elif mode == 3:
        deletes = []
    # mode 1 = 그대로

    total_ops = len(updates) + len(inserts) + len(deletes)
    if total_ops == 0:
        print('  → 적용할 변경 없음.')
        return

    if not auto_yes:
        print()
        ans = input(f'  [{sheet}] 적용하시겠습니까? (y/N): ').strip().lower()
        if ans not in ('y', 'yes'):
            print('  → 건너뜀.')
            return

    print(f'  ▶ 적용 중...')
    u = apply_updates(table, updates) if updates else 0
    i = apply_inserts(table, inserts) if inserts else 0
    d = apply_deletes(table, deletes, pk) if deletes else 0
    print(f'  ✓ UPDATE {u} / INSERT {i} / DELETE {d}')


# ============================================================
# main
# ============================================================
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx', help='업로드할 엑셀 파일 경로')
    ap.add_argument('--mode', type=int, choices=[1, 2, 3], help='1=덮어쓰기(⚠️) / 2=UPDATE만 / 3=차이미리보기(추천)')
    ap.add_argument('--yes', action='store_true', help='확인 없이 적용 (스크립트용)')
    args = ap.parse_args()

    if not os.path.exists(args.xlsx):
        print(f'[ERR] 파일 없음: {args.xlsx}')
        sys.exit(1)

    print(f'엑셀: {args.xlsx}')

    mode = args.mode
    if mode is None:
        print()
        print('=' * 60)
        print(' 업로드 모드 선택')
        print('=' * 60)
        print('  3) 차이 미리보기 후 확인 → 적용 (UPDATE+INSERT)   ⭐⭐ 가장 추천')
        print('  2) id 일치 행 UPDATE만 (신규/삭제 무시)            ⭐ 추천')
        print('  1) 전체 덮어쓰기 (엑셀에 없는 행 = DB 삭제)         ⚠️  비추천')
        print()
        ans = input('  모드 [3]: ').strip() or '3'
        try:
            mode = int(ans)
        except ValueError:
            mode = 3
        if mode not in (1, 2, 3):
            mode = 3

    if mode == 1:
        print()
        print('  ⚠️  WARNING: 모드 1은 엑셀에 없는 모든 행을 DB에서 삭제합니다.')
        print('  ⚠️  되돌리기 어렵습니다. 백업했는지 확인하세요.')
        if not args.yes:
            ans = input('  정말 진행? "DELETE" 입력: ').strip()
            if ans != 'DELETE':
                print('  → 취소됨.')
                sys.exit(0)

    print()
    print(f'=== 모드 {mode} 실행 ===')
    wb = load_workbook(args.xlsx, data_only=True)
    sheets_in_file = wb.sheetnames
    print(f'엑셀 시트: {sheets_in_file}')

    # 처리 순서: Directory → Members → Relations → Ministries
    # (Members 가 household_id, spouse_id 를 참조하므로 households 가 먼저)
    order = ['Directory', 'Members', 'Relations', 'Ministries']
    for sheet in order:
        if sheet not in sheets_in_file:
            continue
        schema = SHEET_SCHEMAS[sheet]
        ws = wb[sheet]
        rows = read_sheet(ws, schema)
        if not rows:
            print(f'\n[{sheet}] 빈 시트 — 건너뜀')
            continue
        run_sheet(sheet, schema, rows, mode, args.yes)

    print()
    print('=' * 60)
    print(' 완료.')
    print('=' * 60)


if __name__ == '__main__':
    main()

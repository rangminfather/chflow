"""
성별 백필 + 가족 관계(부부/부모-자녀) 초기화
- 엑셀 파일에서 부부 row 순서를 참조 (엑셀: 상위행=남편, 하위행=아내)
- 성별 휴리스틱 → members.gender UPDATE
- 같은 household 내 부부 성인 → spouse 관계 (양방향 1쌍)
- 같은 household 내 부부 성인 ↔ 자녀(is_child=true) → parent 관계
"""
import sys, io, os, json, urllib.request, urllib.parse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from collections import defaultdict
from openpyxl import load_workbook

EXCEL = 'c:/csh/project/chflow/test-excel/test_2_fixed_v2.xlsx'
ENV   = 'c:/csh/project/chflow/chflow-app/.env.local'
with open(ENV,'r',encoding='utf-8') as f:
    for line in f:
        line=line.strip()
        if '=' in line and not line.startswith('#'):
            k,v=line.split('=',1); os.environ[k]=v
URL = os.environ['NEXT_PUBLIC_SUPABASE_URL']
KEY = os.environ['SUPABASE_SERVICE_ROLE_KEY']
H = {'apikey':KEY,'Authorization':f'Bearer {KEY}','Content-Type':'application/json','Prefer':'count=exact'}


def _enc(p):
    if '?' in p:
        b,q=p.split('?',1); return b+'?'+urllib.parse.quote(q, safe='=&.,%*!')
    return p

def get_all(path):
    """PostgREST range paging으로 전체 fetch (max_rows=1000 우회)"""
    out = []
    step = 1000
    off = 0
    while True:
        sep = '&' if '?' in path else '?'
        url = f'{URL}/rest/v1{_enc(path + sep + f"offset={off}&limit={step}")}'
        req = urllib.request.Request(url, headers=H)
        data = json.loads(urllib.request.urlopen(req, timeout=60).read())
        out.extend(data)
        if len(data) < step:
            break
        off += step
    return out


# ── 1. DB에서 members + households 로드 ──
print('DB에서 members 로드 중...')
members = get_all('/members?select=id,name,phone,family_church,sub_role,spouse_name,is_child,household_id,source_page,gender')
print(f'  {len(members)}건')

households = get_all('/households?select=id,pasture_id,address,home_phone,order_no')
print(f'households: {len(households)}건')

# 이름+household → member id 빠른 조회
by_hh_name = defaultdict(list)  # (household_id, name) → [member,...]
for m in members:
    by_hh_name[(m['household_id'], m['name'])].append(m)


# ── 2. 엑셀 파싱 → household별 성인 row 순서 ──
print('\n엑셀 파싱 중...')
wb = load_workbook(EXCEL)
ws = wb.active
rows = list(ws.iter_rows(min_row=2, values_only=True))

# family key → [(row_idx, name, phone, sub_role, family_church, spouse, children), ...]
FAM_ROWS = defaultdict(list)
for i, r in enumerate(rows):
    plain, mission, grass, pasture = r[1], r[2], r[3], r[4]
    if not plain or not grass or not pasture:
        continue
    fkey = (plain, grass, pasture, r[11] or '', r[13] or '', r[15] or '')
    name = (r[6] or '').strip()
    if not name:
        continue
    FAM_ROWS[fkey].append({
        'idx': i, 'name': name, 'phone': (r[16] or '').strip(),
        'sub_role': r[9] or '', 'family_church': r[8] or '목원',
        'spouse': (r[7] or '').strip(),
        'children': [c.strip() for c in (r[17] or '').split(',') if c.strip()],
    })
print(f'  가족: {len(FAM_ROWS)}, 엑셀 성인 row: {sum(len(v) for v in FAM_ROWS.values())}')


# ── 3. 성별 추론 ──
# 엑셀의 가족 내 성인 row 순서대로: 첫 행 후보=M, 둘째 행 후보=F (부부 2인 경우)
FEMALE_ROLES = {'권사','은퇴권사','명예권사','원로권사','사모'}
MALE_ROLES   = {'장로','은퇴장로','원로장로','시무장로','시무집사','은퇴시무집사','목사'}
# 전도사, 집사, 명예집사, 청년, 은퇴집사 등은 애매 → 다른 신호

# 이름 끝글자 사전 (간략)
NAME_F = set('숙순희자란혜옥녀연례인임옥엽애이선경은미자연혜정희주하녀원화연')
NAME_M = set('철호훈석규환섭범국준진택용근성식균수환영광기태승섭택탁일오진철호우성필종철민수환수원')
# 겹치는 글자는 위 두 set에 다 있어도 됨 — M 먼저 체크 아님. 아래 규칙 순위로 처리.
NAME_STRONG_F = set('숙순희자란혜옥녀임엽애례녀화')
NAME_STRONG_M = set('철훈석규환섭범국택탁환필종수남근철호준훈승식균광')


def gender_by_role(sub_role, family_church):
    if sub_role in FEMALE_ROLES: return 'F'
    if sub_role in MALE_ROLES:   return 'M'
    if family_church == '목자':  return 'M'
    if family_church == '목녀':  return 'F'
    return None

def gender_by_name(name):
    """이름 끝글자로 강한 신호만 반환"""
    if not name or len(name) < 2:
        return None
    last = name[-1]
    if last in NAME_STRONG_F: return 'F'
    if last in NAME_STRONG_M: return 'M'
    return None


# id → gender 결정
gender_map = {}  # member_id → 'M'/'F'
spouse_pairs = []  # (subject_id, relative_id) 한 쌍 (양쪽 다 insert 할 예정)
parent_links = []  # (child_id, parent_id, role)

stats = {'by_role':0,'by_spouse':0,'by_order':0,'by_name':0,'by_default':0,'unknown':0}

for fkey, frows in FAM_ROWS.items():
    # 각 엑셀 row → DB member id 찾기 (household 단위로)
    # 해당 fkey의 household 찾기: 부부 중 첫 사람의 (name, phone)으로 household_id 조회
    # 단순화: frows[0].name으로 members 중 같은 가족 구성원 household 필터
    # members 엑셀 row와 DB member 매칭:
    hh_candidates = set()
    for fr in frows:
        for m in members:
            if m['name']==fr['name'] and not m['is_child'] and m['phone']==fr['phone']:
                hh_candidates.add(m['household_id'])
                break
    if not hh_candidates:
        continue
    # 한 fkey는 한 household에 대응 (가족 grouping key와 household 생성이 1:1)
    hh_id = next(iter(hh_candidates))

    # 이 household 내 성인 member 목록
    adult_members = [m for m in members if m['household_id']==hh_id and not m['is_child']]
    child_members = [m for m in members if m['household_id']==hh_id and m['is_child']]

    # 엑셀 row 순서대로 matching → DB member id 확정
    row_to_member = []
    for fr in frows:
        mm = next((m for m in adult_members
                   if m['name']==fr['name'] and m['phone']==fr['phone']), None)
        if not mm:
            mm = next((m for m in adult_members if m['name']==fr['name']), None)
        if mm:
            row_to_member.append((fr, mm))

    # 1차: 직분/가정교회 기반
    resolved = {}
    for fr, mm in row_to_member:
        g = gender_by_role(fr['sub_role'], fr['family_church'])
        if g:
            resolved[mm['id']] = (g, 'by_role')

    # 2차: 배우자 반대 성별 (부부 2인 케이스)
    if len(row_to_member) == 2:
        (fr_a, ma), (fr_b, mb) = row_to_member
        if ma['id'] in resolved and mb['id'] not in resolved:
            opp = 'F' if resolved[ma['id']][0]=='M' else 'M'
            resolved[mb['id']] = (opp, 'by_spouse')
        elif mb['id'] in resolved and ma['id'] not in resolved:
            opp = 'F' if resolved[mb['id']][0]=='M' else 'M'
            resolved[ma['id']] = (opp, 'by_spouse')

    # 3차: 부부 2인 둘 다 미확정 → 엑셀 row 순서로 상=M, 하=F
    if len(row_to_member) == 2:
        (fr_a, ma), (fr_b, mb) = row_to_member
        if ma['id'] not in resolved and mb['id'] not in resolved:
            # 엑셀 row idx가 빠른 쪽=M
            if fr_a['idx'] < fr_b['idx']:
                resolved[ma['id']] = ('M','by_order')
                resolved[mb['id']] = ('F','by_order')
            else:
                resolved[mb['id']] = ('M','by_order')
                resolved[ma['id']] = ('F','by_order')

    # 4차: 이름 끝글자
    for fr, mm in row_to_member:
        if mm['id'] not in resolved:
            g = gender_by_name(fr['name'])
            if g:
                resolved[mm['id']] = (g, 'by_name')

    # 5차: default (솔로이면 F)
    for fr, mm in row_to_member:
        if mm['id'] not in resolved:
            resolved[mm['id']] = ('F','by_default')

    # stats
    for mid, (g, src) in resolved.items():
        gender_map[mid] = g
        stats[src] = stats.get(src,0)+1

    # 부부 관계 (성인 2인 이상)
    if len(row_to_member) >= 2:
        ids = [m['id'] for _, m in row_to_member]
        # 단순화: 2명이면 한 쌍, 3명 이상은 pairwise 생략 (데이터상 없음)
        if len(ids) == 2:
            a, b = ids
            spouse_pairs.append((a, b))

    # 부모-자녀 관계
    # 부모 후보: 성인 2명. 성별 확정된 사람을 father/mother로, unknown이면 role NULL.
    fathers = [m['id'] for _, m in row_to_member if gender_map.get(m['id'])=='M']
    mothers = [m['id'] for _, m in row_to_member if gender_map.get(m['id'])=='F']
    for c in child_members:
        for fid in fathers:
            parent_links.append((c['id'], fid, 'father'))
        for mid in mothers:
            parent_links.append((c['id'], mid, 'mother'))


unresolved = [m['id'] for m in members if not m['is_child'] and m['id'] not in gender_map]
stats['unknown'] = len(unresolved)
print(f'\n성별 추론 통계: {stats}')
print(f'자녀(gender=NULL 유지): {sum(1 for m in members if m["is_child"])}')
print(f'부부 쌍: {len(spouse_pairs)}')
print(f'부모-자녀 엣지: {len(parent_links)}')


# ── 4. SQL 생성 ──
print('\nSQL 파일 생성...')
sql_lines = ['BEGIN;']

# gender UPDATE: 성별별로 한 UPDATE in IDs
by_g = defaultdict(list)
for mid, g in gender_map.items():
    by_g[g].append(mid)
for g, ids in by_g.items():
    # 1000개씩 청크
    for i in range(0, len(ids), 500):
        chunk = ids[i:i+500]
        id_list = ",".join(f"'{x}'" for x in chunk)
        sql_lines.append(f"UPDATE public.members SET gender = '{g}' WHERE id IN ({id_list});")

# spouse 관계 (양방향 2쌍)
sql_lines.append('-- spouses (양방향)')
for a, b in spouse_pairs:
    # role: 남편/아내 — 성별로 결정
    ga, gb = gender_map.get(a), gender_map.get(b)
    role_a = 'husband' if ga=='M' else ('wife' if ga=='F' else None)
    role_b = 'husband' if gb=='M' else ('wife' if gb=='F' else None)
    r_a = f"'{role_a}'" if role_a else 'NULL'
    r_b = f"'{role_b}'" if role_b else 'NULL'
    sql_lines.append(
        f"INSERT INTO public.member_relations(subject_id,relative_id,kind,role) "
        f"VALUES ('{a}','{b}','spouse',{r_b}) ON CONFLICT DO NOTHING;"
    )
    sql_lines.append(
        f"INSERT INTO public.member_relations(subject_id,relative_id,kind,role) "
        f"VALUES ('{b}','{a}','spouse',{r_a}) ON CONFLICT DO NOTHING;"
    )

# parent 관계
sql_lines.append('-- parents')
for child, parent, role in parent_links:
    sql_lines.append(
        f"INSERT INTO public.member_relations(subject_id,relative_id,kind,role) "
        f"VALUES ('{child}','{parent}','parent','{role}') ON CONFLICT DO NOTHING;"
    )

sql_lines.append('COMMIT;')

sql_path = 'c:/csh/project/chflow/MS_AX/_backfill_relations.sql'
with open(sql_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql_lines))
print(f'SQL 작성: {sql_path}, {len(sql_lines)} 줄')

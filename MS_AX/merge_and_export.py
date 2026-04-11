"""
사진 페이지 + 목장 페이지 데이터를 cross-reference 매칭
검수용 엑셀 생성 (의심 항목 빨간색)
"""
import sys, io, json, os, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

OUT_DIR = 'C:/csh/project/chflow/MS_AX/parsed-data'
RAW = os.path.join(OUT_DIR, 'raw_data.json')

with open(RAW, 'r', encoding='utf-8') as f:
    raw = json.load(f)

photo_members = raw['photo_members']
pasture_pages = raw['pasture_pages']

# ============================================================
# 매칭: 이름 + 휴대폰 끝 4자리
# ============================================================
def phone_key(phone):
    if not phone:
        return None
    digits = re.sub(r'[^\d]', '', phone)
    return digits[-4:] if len(digits) >= 4 else None


# 사진 페이지 회원: 다중 인덱스 (이름, 휴대폰, 이름+휴대폰)
photo_by_name_phone = {}   # (이름, 전화4자리) → photo_member
photo_by_name = {}          # 이름 → [photo_members]
photo_by_phone = {}         # 전화4자리 → [photo_members]

# 무시할 헤더 단어
HEADER_WORDS = {'번호', '성명', '직분', '주소', '가족', '자택', '휴대폰', '목자', '목녀'}

for pm in photo_members:
    if not pm['name'] or pm['name'] in HEADER_WORDS:
        continue
    photo_by_name.setdefault(pm['name'], []).append(pm)
    pk = phone_key(pm['mobile'])
    if pk:
        photo_by_phone.setdefault(pk, []).append(pm)
        photo_by_name_phone[(pm['name'], pk)] = pm

# 목장 페이지 회원 (가족 단위로 펼쳐서 개인 단위로 변환)
pasture_individuals = []
for page in pasture_pages:
    page_pasture_label = page['shepherd'] + '목장' if page['shepherd'] else None

    # 목자/목녀 자체도 개인으로 추가
    if page['shepherd']:
        pasture_individuals.append({
            'name': page['shepherd'],
            'role': '목자',
            'spouse': page['shepherdess'],
            'home_phone': None,
            'mobile': page['shepherd_phone'],
            'address': '',
            'children': [],
            'page': page['page'],
            'pasture': page_pasture_label,
            'is_shepherd': True,
            'household_id': None,
        })
    if page['shepherdess']:
        pasture_individuals.append({
            'name': page['shepherdess'],
            'role': '목녀',
            'spouse': page['shepherd'],
            'home_phone': None,
            'mobile': page['shepherdess_phone'],
            'address': '',
            'children': [],
            'page': page['page'],
            'pasture': page_pasture_label,
            'is_shepherd': True,
            'household_id': None,
        })

    for h_idx, h in enumerate(page['households']):
        household_id = f"p{page['page']}_h{h['no']}"
        # 헤더 단어 제거
        names = [n for n in h['names'] if n not in HEADER_WORDS]
        roles = h['roles']
        mobiles = h['phone_mobiles']
        for i, name in enumerate(names):
            pasture_individuals.append({
                'name': name,
                'role': roles[i] if i < len(roles) else None,
                'spouse': names[1 - i] if len(names) > 1 else None,
                'home_phone': h['phone_home'],
                'mobile': mobiles[i] if i < len(mobiles) else None,
                'address': h['address'],
                'children': h['children'] if i == 0 else [],  # 자녀는 첫 번째 부모에만
                'page': h['page'],
                'pasture': page_pasture_label,
                'is_shepherd': False,
                'household_id': household_id,
            })

print(f'사진 페이지 회원: {len(photo_members)}')
print(f'목장 페이지 개인: {len(pasture_individuals)}')

# ============================================================
# 매칭 (다단계)
# 1. 이름 + 휴대폰 동시
# 2. 이름 단독 (1명만 매칭되는 경우)
# 3. 휴대폰 단독 (이름 OCR 오류 가정)
# ============================================================
matched = []
used_photo_ids = set()

def try_match(pi):
    """목장 개인을 사진 페이지에 매칭 시도"""
    name = pi['name']
    pk = phone_key(pi.get('mobile'))

    # 1. 이름+휴대폰 정확 매칭
    if name and pk:
        pm = photo_by_name_phone.get((name, pk))
        if pm and id(pm) not in used_photo_ids:
            return pm, 'exact'

    # 2. 이름 단독 매칭 (유일하면 OK)
    if name:
        candidates = photo_by_name.get(name, [])
        avail = [pm for pm in candidates if id(pm) not in used_photo_ids]
        if len(avail) == 1:
            return avail[0], 'name'
        elif len(avail) > 1:
            # 부부 이름으로 좁히기
            spouse = pi.get('spouse')
            if spouse:
                for pm in avail:
                    if pm.get('spouse') == spouse:
                        return pm, 'name+spouse'
            # 첫 번째 후보 반환 (낮은 신뢰도)
            return avail[0], 'name_ambiguous'

    # 3. 휴대폰 단독 매칭
    if pk:
        candidates = photo_by_phone.get(pk, [])
        avail = [pm for pm in candidates if id(pm) not in used_photo_ids]
        if avail:
            return avail[0], 'phone'

    return None, None


unmatched_pasture = []
for pi in pasture_individuals:
    pm, match_type = try_match(pi)

    if pm:
        used_photo_ids.add(id(pm))
        # 매칭 성공!
        matched.append({
            'name': pi['name'],
            'spouse': pi['spouse'] or pm['spouse'],
            'role': pi['role'],
            'home_phone': pi['home_phone'] or pm['home_phone'],
            'mobile': pi['mobile'] or pm['mobile'],
            'plain': pm['plain'],
            'grassland': pm['grassland'],
            'pasture': pi['pasture'] or pm['pasture'],
            'address': pi['address'],
            'children': pi['children'],
            'page_pasture': pi['page'],
            'page_photo': pm['page'],
            'is_shepherd': pi['is_shepherd'],
            'household_id': pi['household_id'],
            'match_status': match_type,
        })
    else:
        unmatched_pasture.append(pi)

# 매칭 안 된 사진 페이지 회원
unmatched_photo = [pm for pm in photo_members
                   if pm['name'] and pm['name'] not in HEADER_WORDS and id(pm) not in used_photo_ids]

print(f'\n매칭 성공: {len(matched)}')
match_types = {}
for m in matched:
    t = m.get('match_status', 'unknown')
    match_types[t] = match_types.get(t, 0) + 1
for t, c in match_types.items():
    print(f'  - {t}: {c}')
print(f'사진만 있음 (목장 페이지에 없음): {len(unmatched_photo)}')
print(f'목장만 있음 (사진 페이지에 없음): {len(unmatched_pasture)}')


# ============================================================
# 엑셀 생성
# ============================================================
wb = openpyxl.Workbook()

# 스타일
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill('solid', fgColor='4F46E5')
SUSPECT_FILL = PatternFill('solid', fgColor='FEE2E2')
MATCHED_FILL = PatternFill('solid', fgColor='F0FDF4')
UNMATCHED_PHOTO_FILL = PatternFill('solid', fgColor='FEF3C7')
UNMATCHED_PASTURE_FILL = PatternFill('solid', fgColor='FEE2E2')
BORDER = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0'),
)


def is_suspect_name(name):
    """OCR 오류 의심 이름"""
    if not name:
        return True
    # 숫자/영문 포함
    if re.search(r'[a-zA-Z0-9]', name):
        return True
    # 한글이 아닌 글자 포함
    if not re.match(r'^[가-힣]+$', name):
        return True
    return False


def is_suspect_phone(phone):
    if not phone:
        return False
    return not re.match(r'^\d{2,3}-\d{3,4}-\d{4}$', phone)


# === Sheet 1: 매칭된 회원 (메인) ===
ws1 = wb.active
ws1.title = '회원 마스터'

headers = ['#', '평원', '초원', '목장', '구분',
           '이름', '배우자', '가정교회', '직분', '주소',
           '자택전화', '휴대폰', '자녀',
           '사진페이지', '목장페이지', '비고']
ws1.append(headers)
for c in range(1, len(headers) + 1):
    cell = ws1.cell(1, c)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = BORDER

row_idx = 2
seq = 1
suspect_count = 0

# 매칭된 것 + 매칭 안 된 사진 페이지 회원 + 매칭 안 된 목장 페이지 회원 모두 출력
all_members = []

# 1. 매칭된 회원 (이미 정확)
for m in matched:
    all_members.append({
        **m,
        'source': 'matched',
    })

# 2. 사진만 있음 (가족 정보 없음)
for pm in unmatched_photo:
    all_members.append({
        'name': pm['name'],
        'spouse': pm['spouse'],
        'role': None,  # 직분 모름
        'home_phone': pm['home_phone'],
        'mobile': pm['mobile'],
        'plain': pm['plain'],
        'grassland': pm['grassland'],
        'pasture': pm['pasture'],
        'address': '',
        'children': [],
        'page_photo': pm['page'],
        'page_pasture': None,
        'household_id': None,
        'is_shepherd': False,
        'source': 'photo_only',
    })

# 3. 목장만 있음 (사진 페이지에 못 찾음)
for pi in unmatched_pasture:
    all_members.append({
        'name': pi['name'],
        'spouse': pi['spouse'],
        'role': pi['role'],
        'home_phone': pi['home_phone'],
        'mobile': pi['mobile'],
        'plain': None,
        'grassland': None,
        'pasture': pi['pasture'],
        'address': pi['address'],
        'children': pi['children'],
        'page_photo': None,
        'page_pasture': pi['page'],
        'household_id': pi['household_id'],
        'is_shepherd': pi['is_shepherd'],
        'source': 'pasture_only',
    })

# 정렬: 평원 > 초원 > 목장 > 가족 순
def sort_key(m):
    return (
        m.get('plain') or 'Z',
        m.get('grassland') or 'Z',
        m.get('pasture') or 'Z',
        m.get('household_id') or 'Z',
    )

all_members.sort(key=sort_key)

def get_family_church(m):
    """가정교회 역할 (목자/목녀/목부/목원) 결정"""
    role = m.get('role') or ''
    if m.get('is_shepherd'):
        # 목자 또는 목녀
        if role in ('목자', '목녀'):
            return role
        # 기본값은 목자 (드문 경우)
        return '목자'
    if role == '목부':
        return '목부'
    return '목원'


def get_ministry_role(m):
    """직분 (시무집사/권사/장로/...) - 가정교회 역할은 제외"""
    role = m.get('role') or ''
    if role in ('목자', '목녀', '목부'):
        return ''  # 가정교회 컬럼으로 이동했으므로 직분은 비움
    return role


for m in all_members:
    family_church = get_family_church(m)
    ministry_role = get_ministry_role(m)

    note = ''
    if m['source'] == 'photo_only':
        note = '⚠ 가족정보없음 (사진페이지만)'
    elif m['source'] == 'pasture_only':
        note = '⚠ 평원/초원정보없음 (목장표만)'

    row_data = [
        seq,
        m.get('plain') or '',
        m.get('grassland') or '',
        m.get('pasture') or '',
        '목자/목녀' if m.get('is_shepherd') else '회원',
        m['name'] or '',
        m.get('spouse') or '',
        family_church,
        ministry_role,
        m.get('address') or '',
        m.get('home_phone') or '',
        m.get('mobile') or '',
        ', '.join(m.get('children') or []),
        m.get('page_photo') or '',
        m.get('page_pasture') or '',
        note,
    ]
    ws1.append(row_data)

    # 의심 항목 빨간색
    is_suspect = (
        is_suspect_name(m['name']) or
        is_suspect_phone(m.get('mobile')) or
        is_suspect_phone(m.get('home_phone')) or
        m['source'] != 'matched'
    )

    fill = None
    if m['source'] == 'photo_only':
        fill = UNMATCHED_PHOTO_FILL
    elif m['source'] == 'pasture_only':
        fill = UNMATCHED_PASTURE_FILL
    elif is_suspect:
        fill = SUSPECT_FILL
    else:
        fill = MATCHED_FILL

    if is_suspect:
        suspect_count += 1

    for c in range(1, len(headers) + 1):
        cell = ws1.cell(row_idx, c)
        cell.fill = fill
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    row_idx += 1
    seq += 1

# 컬럼 너비
widths = [5, 8, 12, 14, 10, 12, 12, 10, 14, 35, 16, 16, 25, 10, 10, 25]
for i, w in enumerate(widths, 1):
    ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# 헤더 행 고정
ws1.freeze_panes = 'A2'

# === Sheet 2: 의심 항목만 ===
ws2 = wb.create_sheet('의심 항목 (검수)')
ws2.append(headers)
for c in range(1, len(headers) + 1):
    cell = ws2.cell(1, c)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = BORDER

# Copy suspect rows from ws1 (using all_members)
seq = 1
for m in all_members:
    is_suspect = (
        is_suspect_name(m['name']) or
        is_suspect_phone(m.get('mobile')) or
        is_suspect_phone(m.get('home_phone')) or
        m['source'] != 'matched'
    )
    if not is_suspect:
        continue

    family_church = get_family_church(m)
    ministry_role = get_ministry_role(m)

    note = ''
    if m['source'] == 'photo_only':
        note = '⚠ 사진페이지만'
    elif m['source'] == 'pasture_only':
        note = '⚠ 목장표만'
    elif is_suspect_name(m['name']):
        note = '⚠ 이름 OCR 의심'

    ws2.append([
        seq,
        m.get('plain') or '',
        m.get('grassland') or '',
        m.get('pasture') or '',
        '목자/목녀' if m.get('is_shepherd') else '회원',
        m['name'] or '',
        m.get('spouse') or '',
        family_church,
        ministry_role,
        m.get('address') or '',
        m.get('home_phone') or '',
        m.get('mobile') or '',
        ', '.join(m.get('children') or []),
        m.get('page_photo') or '',
        m.get('page_pasture') or '',
        note,
    ])
    seq += 1

for i, w in enumerate(widths, 1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws2.freeze_panes = 'A2'

# === Sheet 3: 평원/초원/목장 통계 ===
ws3 = wb.create_sheet('평원·초원·목장 목록')
ws3.append(['평원', '초원', '목장', '회원 수'])
for c in range(1, 5):
    cell = ws3.cell(1, c)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL

stats = defaultdict(int)
for m in all_members:
    key = (m.get('plain'), m.get('grassland'), m.get('pasture'))
    stats[key] += 1

for (p, g, ps), cnt in sorted(stats.items(), key=lambda x: (x[0][0] or 'Z', x[0][1] or 'Z', x[0][2] or 'Z')):
    ws3.append([p or '', g or '', ps or '', cnt])

for i, w in enumerate([10, 16, 18, 10], 1):
    ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# === Sheet 4: 요약 ===
ws4 = wb.create_sheet('요약')
ws4.append(['항목', '값'])
ws4['A1'].font = HEADER_FONT
ws4['A1'].fill = HEADER_FILL
ws4['B1'].font = HEADER_FONT
ws4['B1'].fill = HEADER_FILL
ws4.append(['전체 회원 수', len(all_members)])
ws4.append(['매칭 성공 (안전)', len(matched)])
ws4.append(['사진만 있음 (가족정보 없음)', len(unmatched_photo)])
ws4.append(['목장만 있음 (평원/초원 없음)', len(unmatched_pasture)])
ws4.append(['의심 항목 (검수 필요)', suspect_count])
ws4.append([])
ws4.append(['평원 수', len(set(m.get('plain') for m in all_members if m.get('plain')))])
ws4.append(['초원 수', len(set(m.get('grassland') for m in all_members if m.get('grassland')))])
ws4.append(['목장 수', len(set(m.get('pasture') for m in all_members if m.get('pasture')))])
ws4.column_dimensions['A'].width = 30
ws4.column_dimensions['B'].width = 15

# Save
out_path = os.path.join(OUT_DIR, '명성교회_회원DB_검수용.xlsx')
try:
    wb.save(out_path)
except PermissionError:
    # 파일이 열려있으면 v2로 저장
    out_path = os.path.join(OUT_DIR, '명성교회_회원DB_검수용_v2.xlsx')
    wb.save(out_path)
    print(f'⚠ 원본 파일이 열려있어 v2로 저장합니다')
print(f'\n💾 Excel saved: {out_path}')
print(f'\n📊 통계:')
print(f'  전체 회원: {len(all_members)}')
print(f'  매칭 성공: {len(matched)}')
print(f'  사진만: {len(unmatched_photo)}')
print(f'  목장만: {len(unmatched_pasture)}')
print(f'  의심 항목: {suspect_count}')

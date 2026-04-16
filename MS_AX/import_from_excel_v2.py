"""
test_2_fixed_v2.xlsx → Supabase 일괄 import
- 기존 plains/grasslands/directory_pastures/households/members 전부 삭제 후 재구성
- mission_area 컬럼이 있으면 저장, 없으면 생략
"""
import sys, io, json, os, re, urllib.request, urllib.error, urllib.parse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from collections import defaultdict
from openpyxl import load_workbook

EXCEL = 'c:/csh/project/chflow/test-excel/test_2_fixed_v2.xlsx'
ENV   = 'c:/csh/project/chflow/chflow-app/.env.local'

# ── 환경변수 로드 ──
with open(ENV, 'r', encoding='utf-8') as f:
    for line in f:
        line = line.strip()
        if '=' in line and not line.startswith('#'):
            k, v = line.split('=', 1); os.environ[k] = v
URL = os.environ['NEXT_PUBLIC_SUPABASE_URL']
KEY = os.environ['SUPABASE_SERVICE_ROLE_KEY']
H = {'apikey': KEY, 'Authorization': f'Bearer {KEY}', 'Content-Type': 'application/json', 'Prefer': 'return=representation'}


def _enc(path):
    if '?' in path:
        base, q = path.split('?', 1)
        return base + '?' + urllib.parse.quote(q, safe='=&.,%*!')
    return path

def http(method, path, body=None):
    url = f'{URL}/rest/v1{_enc(path)}'
    data = json.dumps(body, ensure_ascii=False).encode('utf-8') if body is not None else None
    req = urllib.request.Request(url, data=data, headers=H, method=method)
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            raw = r.read()
            return json.loads(raw) if raw else []
    except urllib.error.HTTPError as e:
        msg = e.read().decode('utf-8', errors='replace')
        print(f'  HTTP {e.code} {method} {path}: {msg[:300]}')
        return None


def insert(table, rows):
    if not rows: return []
    out = []
    for i in range(0, len(rows), 100):
        res = http('POST', f'/{table}', rows[i:i+100])
        if res: out.extend(res)
    return out


def delete_all(table):
    http('DELETE', f'/{table}?id=neq.00000000-0000-0000-0000-000000000000')


# ── mission_area 컬럼 존재 여부 ──
probe = http('GET', '/directory_pastures?select=mission_area&limit=1')
HAS_MISSION = probe is not None
print(f'mission_area 컬럼: {"있음" if HAS_MISSION else "없음 (import 후 migration 실행 후 재실행 권장)"}')


# ── 엑셀 파싱 ──
print('\n=== 엑셀 파싱 ===')
wb = load_workbook(EXCEL)
ws = wb.active

rows = list(ws.iter_rows(min_row=2, values_only=True))
print(f'총 행: {len(rows)}')

def norm_name(v):
    return str(v).strip() if v is not None else ''

def norm_phone(v):
    return str(v).strip() if v is not None else ''

def parse_children(v):
    if not v: return []
    return [n.strip() for n in str(v).split(',') if n.strip()]

def parse_birth(v):
    if not v: return None
    s = str(v).strip()
    # YYYY-MM-DD, YYYY.MM.DD, YYYYMMDD 등 단순 처리
    m = re.match(r'^(\d{4})[.\-/]?(\d{1,2})[.\-/]?(\d{1,2})$', s)
    if m:
        y, mo, d = m.groups()
        return f'{y}-{int(mo):02d}-{int(d):02d}'
    return None


# ── 가족 그룹핑 ──
# 가족 키: (평원, 초원, 목장, 도로명, 상세, 자택)
FAMILIES = defaultdict(list)    # key → [row,...]
FAMILY_ORDER = {}               # key → 첫 등장 # 번호
for r in rows:
    plain, mission, grass, pasture = r[1], r[2], r[3], r[4]
    if not plain or not grass or not pasture:
        continue
    fkey = (plain, grass, pasture, r[11] or '', r[13] or '', r[15] or '')
    FAMILIES[fkey].append(r)
    if fkey not in FAMILY_ORDER:
        FAMILY_ORDER[fkey] = r[0] or 0
print(f'가족: {len(FAMILIES)}')

# ── (plain, grass, pasture) → mission_area ──
pasture_mission = {}
for r in rows:
    if r[1] and r[3] and r[4]:
        pasture_mission[(r[1], r[3], r[4])] = r[2]

# ── 평원/초원/목장 세트 ──
plains_set = sorted({r[1] for r in rows if r[1]},
                    key=lambda x: (0, int(x)) if str(x).isdigit() else (1, str(x)))
grasslands_set = {}  # (plain, grass) → order_no
pastures_set   = {}  # (plain, grass, pasture) → order_no
for r in rows:
    if r[1] and r[3]:
        grasslands_set.setdefault((r[1], r[3]), r[0] or 0)
    if r[1] and r[3] and r[4]:
        pastures_set.setdefault((r[1], r[3], r[4]), r[0] or 0)


# ── 기존 데이터 삭제 (FK 순서) ──
print('\n=== 기존 데이터 삭제 ===')
print('  members 삭제...');           delete_all('members')
print('  households 삭제...');        delete_all('households')
print('  directory_pastures 삭제...'); delete_all('directory_pastures')
print('  grasslands 삭제...');        delete_all('grasslands')
print('  plains 삭제...');            delete_all('plains')


# ── plains ──
print('\n=== plains ===')
plain_rows = [
    {'name': str(p),
     'display_name': f'{p}평원' if p != '젊은이' else '젊은이평원',
     'order_no': i}
    for i, p in enumerate(plains_set)
]
inserted = insert('plains', plain_rows)
plain_id = {r['name']: r['id'] for r in inserted}
print(f'  {len(inserted)}건')

# ── grasslands ──
print('\n=== grasslands ===')
grass_rows = [
    {'plain_id': plain_id[str(p)], 'name': g, 'order_no': 0}
    for (p, g), _ in grasslands_set.items()
]
inserted = insert('grasslands', grass_rows)
grass_id = {}
for r in inserted:
    # plain 이름 역조회
    for pname, pid in plain_id.items():
        if pid == r['plain_id']:
            grass_id[(pname, r['name'])] = r['id']
            break
print(f'  {len(inserted)}건')

# ── directory_pastures ──
print('\n=== directory_pastures ===')
past_rows = []
for (p, g, ps), _ in pastures_set.items():
    gid = grass_id.get((str(p), g))
    if not gid: continue
    row = {'grassland_id': gid, 'name': ps, 'order_no': 0}
    if HAS_MISSION:
        row['mission_area'] = pasture_mission.get((p, g, ps))
    past_rows.append(row)
inserted = insert('directory_pastures', past_rows)
past_id = {}
# (plain,grass,pasture) → id 매핑 재구축
for r in inserted:
    # grass_id → (plain,grass) 역조회
    for (pn, gn), gid in grass_id.items():
        if gid == r['grassland_id']:
            past_id[(pn, gn, r['name'])] = r['id']
            break
print(f'  {len(inserted)}건')


# ── households + members ──
print('\n=== households + members ===')
households_rows = []
family_order_map = []  # households 순서에 대응하는 family key
for fkey, frows in FAMILIES.items():
    plain, grass, pasture, road, detail, home_phone = fkey
    pid = past_id.get((str(plain), grass, pasture))
    if not pid: continue
    # 주소는 도로명 + 상세주소 합쳐서
    addr_parts = [x for x in [road, detail] if x]
    households_rows.append({
        'pasture_id': pid,
        'address': ' '.join(addr_parts),
        'home_phone': home_phone,
        'order_no': FAMILY_ORDER[fkey],
    })
    family_order_map.append(fkey)

inserted_h = insert('households', households_rows)
print(f'  households: {len(inserted_h)}건')

# ── members ──
# guard_status는 테이블 default('비회원')에 맡긴다 (link_member_to_signup RPC가 가입 시 '회원'으로 전환)
DEFAULT_MEMBER = {
    'name': '', 'phone': '', 'birth_date': None,
    'household_id': None, 'family_church': '목원', 'sub_role': '',
    'spouse_name': '',
    'source_page': None, 'is_child': False, 'notes': None,
}
members_rows = []
for idx, fkey in enumerate(family_order_map):
    if idx >= len(inserted_h): break
    h_id = inserted_h[idx]['id']
    frows = FAMILIES[fkey]
    children_added = set()
    for r in frows:
        name = norm_name(r[6])
        if not name: continue
        m = {**DEFAULT_MEMBER,
             'name': name,
             'phone': norm_phone(r[16]),
             'birth_date': parse_birth(r[10]),
             'household_id': h_id,
             'family_church': r[8] or '목원',
             'sub_role': r[9] or '',
             'spouse_name': norm_name(r[7]),
             'source_page': r[19] if isinstance(r[19], int) else None,
             'is_child': False,
             'notes': str(r[20]) if r[20] else None}
        members_rows.append(m)
        # 자녀
        for c in parse_children(r[17]):
            if c in children_added: continue
            children_added.add(c)
            members_rows.append({**DEFAULT_MEMBER,
                'name': c,
                'household_id': h_id,
                'family_church': '목원',
                'source_page': r[19] if isinstance(r[19], int) else None,
                'is_child': True})

inserted_m = insert('members', members_rows)
print(f'  members: {len(inserted_m)}건')

# ── 통계 ──
print('\n=== 완료 ===')
print(f'plains:             {len(plain_id)}')
print(f'grasslands:         {len(grass_id)}')
print(f'directory_pastures: {len(past_id)}')
print(f'households:         {len(inserted_h)}')
print(f'members:            {len(inserted_m)}')

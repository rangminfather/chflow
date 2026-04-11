"""
V3 매칭 + 엑셀 생성
- 부부 단위 매칭 추가
- 자녀 별도 row
- 사용자 수정본 형식 매칭
"""
import sys, io, json, os, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

OUT_DIR = 'C:/csh/project/chflow/MS_AX/parsed-data'
RAW = os.path.join(OUT_DIR, 'raw_data_v3.json')

with open(RAW, 'r', encoding='utf-8') as f:
    raw = json.load(f)

photo_members = raw['photo_members']
pasture_pages = raw['pasture_pages']

HEADER_WORDS = {'번호', '성명', '직분', '주소', '가족', '자택', '휴대폰', '목자', '목녀'}


def phone_key(phone):
    if not phone:
        return None
    digits = re.sub(r'[^\d]', '', phone)
    return digits[-4:] if len(digits) >= 4 else None


# ============================================================
# 사진 페이지 인덱싱 (다단계)
# ============================================================
photo_by_name = defaultdict(list)
photo_by_phone = defaultdict(list)
photo_by_name_phone = {}

for pm in photo_members:
    if not pm['name'] or pm['name'] in HEADER_WORDS:
        continue
    photo_by_name[pm['name']].append(pm)
    pk = phone_key(pm['mobile'])
    if pk:
        photo_by_phone[pk].append(pm)
        photo_by_name_phone[(pm['name'], pk)] = pm

# 휴대폰 끝 4자리 → 사진 회원 (역방향)
phone_to_photo = defaultdict(list)
for pm in photo_members:
    pk = phone_key(pm['mobile'])
    if pk:
        phone_to_photo[pk].append(pm)

# 사진 회원을 목장 이름으로 인덱싱
photo_by_pasture_name = defaultdict(list)
for pm in photo_members:
    if pm['pasture']:
        photo_by_pasture_name[pm['pasture']].append(pm)


# ============================================================
# 목장 페이지 → 가족 리스트 만들기
# ============================================================
households_list = []  # 가족 단위
member_seq = 0

for page in pasture_pages:
    page_pasture = page['shepherd']  # 목장 이름 = 목자 이름
    if not page_pasture:
        continue

    # 목자/목녀 자체를 가족 0으로 추가
    leader_household = {
        'page': page['page'],
        'no': 0,
        'pasture_name': page_pasture,
        'is_leader': True,
        'members': [],
        'address': '',
        'children': [],
        'phone_home': None,
    }
    if page['shepherd']:
        leader_household['members'].append({
            'name': page['shepherd'],
            'role': '목자',
            'family_church': '목자',
            'mobile': page['shepherd_phone'],
            'is_shepherd': True,
        })
    if page['shepherdess']:
        leader_household['members'].append({
            'name': page['shepherdess'],
            'role': '목녀',
            'family_church': '목녀',
            'mobile': page['shepherdess_phone'],
            'is_shepherd': True,
        })
    if leader_household['members']:
        households_list.append(leader_household)

    # 일반 가족
    for h in page['households']:
        names = [n for n in h['names'] if n not in HEADER_WORDS]
        if not names:
            continue
        roles = h['roles']
        mobiles = h['phone_mobiles']

        members = []
        for i, name in enumerate(names):
            members.append({
                'name': name,
                'role': roles[i] if i < len(roles) else None,
                'family_church': '목원',
                'mobile': mobiles[i] if i < len(mobiles) else None,
                'is_shepherd': False,
            })

        households_list.append({
            'page': h['page'],
            'no': h['no'],
            'pasture_name': page_pasture,
            'is_leader': False,
            'members': members,
            'address': h.get('address', ''),
            'children': h.get('children', []),
            'phone_home': h.get('phone_home'),
        })


# ============================================================
# 매칭: 가족 단위 매칭 우선, 그 다음 개별 매칭
# ============================================================
used_photo_ids = set()
matched_households = []

def match_member(name, mobile, pasture_name=None):
    """개인 매칭 - 같은 목장 내 우선"""
    pk = phone_key(mobile)

    # Pass 1: 같은 목장 + 이름 정확
    if name and pasture_name:
        for pm in photo_by_pasture_name.get(pasture_name, []):
            if pm['name'] == name and id(pm) not in used_photo_ids:
                return pm

    # Pass 2: 같은 목장 + 휴대폰 끝 4자리
    if pk and pasture_name:
        for pm in photo_by_pasture_name.get(pasture_name, []):
            if phone_key(pm['mobile']) == pk and id(pm) not in used_photo_ids:
                return pm

    # Pass 3: 이름 + 휴대폰 정확 (목장 무관)
    if name and pk:
        pm = photo_by_name_phone.get((name, pk))
        if pm and id(pm) not in used_photo_ids:
            return pm

    # Pass 4: 이름 단독 (1명 후보면 안전)
    if name:
        avail = [pm for pm in photo_by_name.get(name, []) if id(pm) not in used_photo_ids]
        if len(avail) == 1:
            return avail[0]
        elif len(avail) > 1:
            # 같은 목장 우선
            if pasture_name:
                same = [pm for pm in avail if pm.get('pasture') == pasture_name]
                if same:
                    return same[0]
            return avail[0]

    # Pass 5: 휴대폰 단독
    if pk:
        avail = [pm for pm in photo_by_phone.get(pk, []) if id(pm) not in used_photo_ids]
        if avail:
            if pasture_name:
                same = [pm for pm in avail if pm.get('pasture') == pasture_name]
                if same:
                    return same[0]
            return avail[0]

    return None


for h in households_list:
    matched_h = {**h, 'matched_members': []}
    for m in h['members']:
        pm = match_member(m['name'], m['mobile'], h['pasture_name'])
        match_info = None
        if pm:
            used_photo_ids.add(id(pm))
            match_info = pm
        matched_h['matched_members'].append({
            **m,
            'photo_match': match_info,
        })
    matched_households.append(matched_h)


# 매칭 안 된 사진 회원 (아직도 photo_only)
unmatched_photo = [pm for pm in photo_members
                   if pm['name'] and pm['name'] not in HEADER_WORDS
                   and id(pm) not in used_photo_ids]


print(f'총 가족: {len(matched_households)}')
total_pasture_members = sum(len(h['members']) for h in matched_households)
total_matched = sum(1 for h in matched_households for mm in h['matched_members'] if mm['photo_match'])
print(f'목장표 개인: {total_pasture_members}')
print(f'매칭 성공: {total_matched}')
print(f'사진만: {len(unmatched_photo)}')


# ============================================================
# 엑셀 생성
# ============================================================
wb = openpyxl.Workbook()

HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill('solid', fgColor='4F46E5')
SUSPECT_FILL = PatternFill('solid', fgColor='FEE2E2')
MATCHED_FILL = PatternFill('solid', fgColor='F0FDF4')
PHOTO_ONLY_FILL = PatternFill('solid', fgColor='FEF3C7')
PASTURE_ONLY_FILL = PatternFill('solid', fgColor='FDE68A')
LEADER_FILL = PatternFill('solid', fgColor='DBEAFE')
BORDER = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0'),
)


def is_suspect_name(name):
    if not name:
        return True
    if re.search(r'[a-zA-Z0-9]', name):
        return True
    if not re.match(r'^[가-힣]+$', name):
        return True
    return False


def is_suspect_phone(phone):
    if not phone:
        return False
    return not re.match(r'^\d{2,3}-\d{3,4}-\d{4}$', phone)


# === Sheet 1: 회원 마스터 ===
ws1 = wb.active
ws1.title = '회원 마스터'

headers = ['#', '평원', '초원', '목장', '구분',
           '이름', '배우자', '가정교회', '직분', '주소',
           '자택전화', '휴대폰', '자녀', '사진페이지', '페이지', '비고']
ws1.append(headers)
for c in range(1, len(headers) + 1):
    cell = ws1.cell(1, c)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = BORDER

row_idx = 2
seq = 1

# 정렬: 평원 → 초원 → 목장
def sort_household(h):
    pm = h['matched_members'][0]['photo_match'] if h['matched_members'] and h['matched_members'][0]['photo_match'] else None
    plain = pm['plain'] if pm else 'Z'
    grass = pm['grassland'] if pm else 'Z'
    return (plain or 'Z', grass or 'Z', h['pasture_name'], h['no'])

matched_households.sort(key=sort_household)


def add_member_row(seq, plain, grassland, pasture, guard, name, spouse,
                   family_church, role, address, home_phone, mobile,
                   children, page_photo, page_pasture, note, fill):
    ws1.append([
        seq, plain or '', grassland or '', pasture or '', guard,
        name or '', spouse or '',
        family_church or '', role or '', address or '',
        home_phone or '', mobile or '',
        children or '',
        page_photo or '', page_pasture or '', note or '',
    ])
    for c in range(1, len(headers) + 1):
        cell = ws1.cell(ws1.max_row, c)
        cell.fill = fill
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)


for h in matched_households:
    members = h['matched_members']
    if not members:
        continue

    # 주소/자택은 가족당 1번만
    addr = h['address']
    home_phone = h['phone_home']
    children_str = ', '.join(h['children'])

    # 사진 페이지 정보 (matched member에서 가져옴)
    photo_info = None
    for mm in members:
        if mm['photo_match']:
            photo_info = mm['photo_match']
            break
    plain = photo_info['plain'] if photo_info else None
    grassland = photo_info['grassland'] if photo_info else None
    pasture = photo_info['pasture'] if photo_info else h['pasture_name']

    # 부부 (1~2명)
    for i, mm in enumerate(members):
        pm = mm['photo_match']
        spouse = members[1 - i]['name'] if len(members) > 1 else None

        is_leader = h['is_leader']
        fill = MATCHED_FILL if pm else (LEADER_FILL if is_leader else PASTURE_ONLY_FILL)
        if is_suspect_name(mm['name']) or is_suspect_phone(mm['mobile']):
            fill = SUSPECT_FILL

        note = ''
        if not pm:
            note = '⚠ 사진페이지 매칭 실패'

        add_member_row(
            seq=seq,
            plain=plain or '',
            grassland=grassland or '',
            pasture=pasture or '',
            guard='비회원',
            name=mm['name'],
            spouse=spouse,
            family_church=mm['family_church'],
            role=mm['role'] if mm['family_church'] == '목원' else '',
            address=addr if i == 0 else '',
            home_phone=home_phone if i == 0 else '',
            mobile=mm['mobile'],
            children=children_str if i == 0 else '',
            page_photo=pm['page'] if pm else '',
            page_pasture=h['page'],
            note=note,
            fill=fill,
        )
        seq += 1

    # 자녀는 별도 row 만들지 않고 가족 컬럼 텍스트로만 유지
    # (사용자 v2.1 형식에 맞춤)


# 매칭 안 된 사진 페이지 회원 (가족 정보 없음)
for pm in unmatched_photo:
    add_member_row(
        seq=seq,
        plain=pm['plain'] or '',
        grassland=pm['grassland'] or '',
        pasture=pm['pasture'] or '',
        guard='비회원',
        name=pm['name'],
        spouse=pm['spouse'],
        family_church='목원',
        role='',
        address='',
        home_phone=pm['home_phone'],
        mobile=pm['mobile'],
        children='',
        page_photo=pm['page'],
        page_pasture='',
        note='⚠ 가족정보없음 (사진페이지만)',
        fill=PHOTO_ONLY_FILL,
    )
    seq += 1


widths = [5, 8, 14, 14, 8, 12, 12, 10, 14, 35, 16, 16, 25, 8, 8, 25]
for i, w in enumerate(widths, 1):
    ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws1.freeze_panes = 'A2'


# === Sheet 2: 평원/초원/목장 통계 ===
ws3 = wb.create_sheet('평원·초원·목장 목록')
ws3.append(['평원', '초원', '목장', '회원 수'])
for c in range(1, 5):
    ws3.cell(1, c).font = HEADER_FONT
    ws3.cell(1, c).fill = HEADER_FILL

stats = defaultdict(int)
for r in range(2, ws1.max_row + 1):
    p = ws1.cell(r, 2).value or '?'
    g = ws1.cell(r, 3).value or '?'
    ps = ws1.cell(r, 4).value or '?'
    stats[(p, g, ps)] += 1

for (p, g, ps), cnt in sorted(stats.items()):
    ws3.append([p, g, ps, cnt])

for i, w in enumerate([10, 18, 18, 10], 1):
    ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


# === Sheet 3: 요약 ===
ws4 = wb.create_sheet('요약')
ws4.append(['항목', '값'])
ws4['A1'].font = HEADER_FONT
ws4['A1'].fill = HEADER_FILL
ws4['B1'].font = HEADER_FONT
ws4['B1'].fill = HEADER_FILL
ws4.append(['전체 row 수', ws1.max_row - 1])
ws4.append(['가족 단위 수', len(matched_households)])
ws4.append(['목장표 개인', total_pasture_members])
ws4.append(['매칭 성공', total_matched])
ws4.append(['사진만', len(unmatched_photo)])
ws4.column_dimensions['A'].width = 25
ws4.column_dimensions['B'].width = 12


# Save
out_path = os.path.join(OUT_DIR, '명성교회_회원DB_v3.xlsx')
try:
    wb.save(out_path)
except PermissionError:
    out_path = os.path.join(OUT_DIR, '명성교회_회원DB_v3_new.xlsx')
    wb.save(out_path)

print(f'\n💾 Saved: {out_path}')
print(f'📊 총 row: {ws1.max_row - 1}')

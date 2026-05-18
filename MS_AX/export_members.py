"""
성도정보 익스포트 → 엑셀 (xlsx)

사용:
  python export_members.py
  python export_members.py --out path/to/file.xlsx
  python export_members.py --all       # 모든 시트 포함 (질문 생략)
  python export_members.py --basic     # Members + Relations만 (질문 생략)

생성 시트:
  - Members      (필수): 성도 기본정보 + 목장경로/주소/계정여부
  - Relations    (선택): 가족 관계 (subject/relative + kind/role)
  - Ministries   (선택): 직분/사역
  - Directory    (선택): 평원/초원/목장/가족 4계층

* 모든 시트 첫 컬럼은 id (UUID) — 업로드 시 매칭에 사용. 절대 수정/삭제 금지.
"""
import sys, io, os, json, urllib.request, urllib.error, urllib.parse
import argparse
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.comments import Comment
except ImportError:
    print('[ERROR] openpyxl 필요: pip install openpyxl')
    sys.exit(1)


# ============================================================
# 환경
# ============================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ENV_PATH = 'c:/csh/project/chflow/chflow-app/.env.local'

with open(ENV_PATH, 'r', encoding='utf-8') as f:
    for line in f:
        line = line.strip()
        if '=' in line and not line.startswith('#'):
            k, v = line.split('=', 1)
            os.environ[k] = v

URL = os.environ['NEXT_PUBLIC_SUPABASE_URL']
KEY = os.environ['SUPABASE_SERVICE_ROLE_KEY']
H = {
    'apikey': KEY,
    'Authorization': f'Bearer {KEY}',
    'Content-Type': 'application/json',
}


def _enc(path):
    if '?' in path:
        base, q = path.split('?', 1)
        return base + '?' + urllib.parse.quote(q, safe='=&.,%*!')
    return path


def fetch_all(table, select='*', extra=''):
    """Supabase REST를 통한 전체 페이지 fetch."""
    rows = []
    page = 0
    PAGE_SIZE = 1000
    while True:
        path = f'/{table}?select={select}'
        if extra:
            path += '&' + extra
        url = f'{URL}/rest/v1{_enc(path)}'
        req = urllib.request.Request(url, headers={
            **H,
            'Range-Unit': 'items',
            'Range': f'{page * PAGE_SIZE}-{(page + 1) * PAGE_SIZE - 1}',
        })
        try:
            with urllib.request.urlopen(req, timeout=120) as r:
                chunk = json.loads(r.read())
        except urllib.error.HTTPError as e:
            print(f'[ERROR] HTTP {e.code} {table}: {e.read().decode("utf-8", errors="replace")[:300]}')
            return []
        if not chunk:
            break
        rows.extend(chunk)
        if len(chunk) < PAGE_SIZE:
            break
        page += 1
    return rows


# ============================================================
# 옵션
# ============================================================
SHEET_DEFS = [
    ('Members',    True,  '성도 기본정보 (필수)'),
    ('Relations',  True,  '가족관계 (부모/자녀/배우자)'),
    ('Ministries', False, '직분/사역 (member_ministries)'),
    ('Directory',  False, '평원/초원/목장/가족 구조'),
]


def ask_sheets(args):
    if args.all:
        return [name for name, _, _ in SHEET_DEFS]
    if args.basic:
        return ['Members', 'Relations']

    print()
    print('=' * 60)
    print(' 익스포트 시트 선택')
    print('=' * 60)
    chosen = []
    for name, default, desc in SHEET_DEFS:
        if name == 'Members':
            print(f'  [✓] {name:<10s} {desc}')
            chosen.append(name)
            continue
        d = 'Y/n' if default else 'y/N'
        ans = input(f'  포함? {name:<10s} ({desc}) [{d}]: ').strip().lower()
        if ans == '':
            include = default
        else:
            include = ans in ('y', 'yes', '예', 'ㅇ')
        if include:
            chosen.append(name)
            print(f'  [✓] {name}')
        else:
            print(f'  [ ] {name}')
    return chosen


def ask_output(args):
    if args.out:
        return args.out
    default = os.path.join(
        SCRIPT_DIR, 'exports',
        f'members_export_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    )
    print()
    ans = input(f'저장 경로 (Enter=기본 {default}): ').strip()
    return ans or default


# ============================================================
# 시트 빌더
# ============================================================
def write_header(ws, headers, frozen_cols=2, key_cols=None):
    key_cols = key_cols or {'id'}
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='305496')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if h in key_cols:
            cell.fill = PatternFill('solid', fgColor='8B4513')
            cell.comment = Comment('PK — 절대 수정/삭제 금지', 'system')
    ws.freeze_panes = ws.cell(row=2, column=frozen_cols + 1).coordinate


def auto_width(ws, headers, sample_rows=200):
    for ci, h in enumerate(headers, 1):
        max_len = len(str(h))
        for r in range(2, min(2 + sample_rows, ws.max_row + 1)):
            v = ws.cell(row=r, column=ci).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max(max_len + 2, 8), 40)


def build_members(wb, members, households, pastures, grasslands, plains):
    h_map = {h['id']: h for h in households}
    p_map = {p['id']: p for p in pastures}
    g_map = {g['id']: g for g in grasslands}
    pl_map = {p['id']: p for p in plains}

    ws = wb.create_sheet('Members')
    headers = [
        'id', 'excel_row_no', 'name', 'gender', 'birth_date', 'phone',
        'family_church', 'sub_role', 'spouse_name',
        'plain_name', 'grassland_name', 'pasture_name', 'address',
        'is_child', 'guard_status', 'has_account',
        'photo_status', 'photo_page', 'photo_url',
        'source_page', 'notes',
        'household_id', 'spouse_id',
    ]
    write_header(ws, headers, frozen_cols=3, key_cols={'id', 'household_id', 'spouse_id'})

    members_sorted = sorted(members, key=lambda m: (m.get('excel_row_no') or 99999, m.get('name') or ''))
    for r, m in enumerate(members_sorted, 2):
        hh = h_map.get(m.get('household_id')) or {}
        past = p_map.get(hh.get('pasture_id')) or {}
        gr = g_map.get(past.get('grassland_id')) or {}
        pl = pl_map.get(gr.get('plain_id')) or {}

        row = [
            m.get('id'),
            m.get('excel_row_no'),
            m.get('name'),
            m.get('gender'),
            m.get('birth_date'),
            m.get('phone'),
            m.get('family_church'),
            m.get('sub_role'),
            m.get('spouse_name'),
            pl.get('name'),
            gr.get('name'),
            past.get('name'),
            hh.get('address'),
            m.get('is_child'),
            m.get('guard_status'),
            'Y' if m.get('app_user_id') else '',
            m.get('photo_status'),
            m.get('photo_page'),
            m.get('photo_url'),
            m.get('source_page'),
            m.get('notes'),
            m.get('household_id'),
            m.get('spouse_id'),
        ]
        for ci, v in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=v)
    auto_width(ws, headers)
    print(f'  [Members]    {len(members_sorted)} 행')


def build_relations(wb, relations, members):
    m_map = {m['id']: m for m in members}
    ws = wb.create_sheet('Relations')
    headers = [
        'id', 'subject_id', 'subject_name',
        'relative_id', 'relative_name',
        'kind', 'role',
    ]
    write_header(ws, headers, frozen_cols=3, key_cols={'id', 'subject_id', 'relative_id'})

    rel_sorted = sorted(relations, key=lambda r: (
        (m_map.get(r.get('subject_id')) or {}).get('name') or '',
        r.get('kind') or '',
    ))
    for r, rel in enumerate(rel_sorted, 2):
        s = m_map.get(rel.get('subject_id')) or {}
        rv = m_map.get(rel.get('relative_id')) or {}
        row = [
            rel.get('id'),
            rel.get('subject_id'),
            s.get('name'),
            rel.get('relative_id'),
            rv.get('name'),
            rel.get('kind'),
            rel.get('role'),
        ]
        for ci, v in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=v)
    auto_width(ws, headers)
    print(f'  [Relations]  {len(rel_sorted)} 행')


def build_ministries(wb, ministries, members):
    m_map = {m['id']: m for m in members}
    ws = wb.create_sheet('Ministries')
    headers = ['id', 'member_id', 'member_name', 'ministry', 'role', 'notes']
    write_header(ws, headers, frozen_cols=3, key_cols={'id', 'member_id'})

    mn_sorted = sorted(ministries, key=lambda x: (
        x.get('ministry') or '',
        (m_map.get(x.get('member_id')) or {}).get('name') or '',
    ))
    for r, x in enumerate(mn_sorted, 2):
        m = m_map.get(x.get('member_id')) or {}
        row = [x.get('id'), x.get('member_id'), m.get('name'), x.get('ministry'), x.get('role'), x.get('notes')]
        for ci, v in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=v)
    auto_width(ws, headers)
    print(f'  [Ministries] {len(mn_sorted)} 행')


def build_directory(wb, plains, grasslands, pastures, households):
    pl_map = {p['id']: p for p in plains}
    g_map = {g['id']: g for g in grasslands}
    p_map = {p['id']: p for p in pastures}

    ws = wb.create_sheet('Directory')
    headers = [
        'household_id', 'plain_name', 'grassland_name', 'pasture_name',
        'address', 'home_phone', 'order_no',
        'pasture_id', 'grassland_id', 'plain_id',
    ]
    write_header(ws, headers, frozen_cols=4,
                 key_cols={'household_id', 'pasture_id', 'grassland_id', 'plain_id'})

    rows = []
    for h in households:
        past = p_map.get(h.get('pasture_id')) or {}
        gr = g_map.get(past.get('grassland_id')) or {}
        pl = pl_map.get(gr.get('plain_id')) or {}
        rows.append((
            h['id'], pl.get('name'), gr.get('name'), past.get('name'),
            h.get('address'), h.get('home_phone'), h.get('order_no'),
            past.get('id'), gr.get('id'), pl.get('id'),
        ))
    rows.sort(key=lambda x: (x[1] or '', x[2] or '', x[3] or '', x[6] or 0))
    for r, row in enumerate(rows, 2):
        for ci, v in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=v)
    auto_width(ws, headers)
    print(f'  [Directory]  {len(rows)} 행')


# ============================================================
# main
# ============================================================
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--out', help='저장 경로 (.xlsx)')
    ap.add_argument('--all', action='store_true', help='모든 시트 포함')
    ap.add_argument('--basic', action='store_true', help='Members+Relations만')
    args = ap.parse_args()

    sheets = ask_sheets(args)
    out_path = ask_output(args)
    os.makedirs(os.path.dirname(out_path) or '.', exist_ok=True)

    print()
    print('=== DB에서 가져오는 중 ===')
    members = fetch_all('members')
    print(f'  members:    {len(members)}')

    households = fetch_all('households')
    pastures = fetch_all('directory_pastures')
    grasslands = fetch_all('grasslands')
    plains = fetch_all('plains')
    print(f'  households: {len(households)} / pastures: {len(pastures)} / '
          f'grasslands: {len(grasslands)} / plains: {len(plains)}')

    relations = []
    if 'Relations' in sheets:
        relations = fetch_all('member_relations')
        print(f'  relations:  {len(relations)}')

    ministries = []
    if 'Ministries' in sheets:
        ministries = fetch_all('member_ministries')
        print(f'  ministries: {len(ministries)}')

    print()
    print('=== 엑셀 생성 ===')
    wb = Workbook()
    wb.remove(wb.active)

    if 'Members' in sheets:
        build_members(wb, members, households, pastures, grasslands, plains)
    if 'Relations' in sheets:
        build_relations(wb, relations, members)
    if 'Ministries' in sheets:
        build_ministries(wb, ministries, members)
    if 'Directory' in sheets:
        build_directory(wb, plains, grasslands, pastures, households)

    info = wb.create_sheet('_README', 0)
    info.append(['chflow 성도정보 익스포트'])
    info.append(['생성일시', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    info.append(['포함시트', ', '.join(sheets)])
    info.append([])
    info.append(['주의사항'])
    info.append(['1. id 컬럼 (갈색 헤더) — 절대 수정/삭제하지 마세요. 업로드 시 매칭 키입니다.'])
    info.append(['2. 새 행 추가 시 id 칸은 비워두면 자동 생성됩니다.'])
    info.append(['3. 행 삭제는 업로드 모드 1(전체 덮어쓰기)에서만 반영됩니다.'])
    info.append(['4. plain_name/grassland_name/pasture_name/address 등 _name 컬럼은 참고용입니다.'])
    info.append(['   소속 변경은 household_id 를 다른 값으로 바꿔야 반영됩니다.'])
    info.append(['5. 업로드: python import_members.py 로 실행하세요.'])
    for col in 'A':
        info.column_dimensions[col].width = 90
    info['A1'].font = Font(bold=True, size=14)

    wb.save(out_path)

    size_kb = os.path.getsize(out_path) / 1024
    print()
    print('=' * 60)
    print(f' ✓ 완료: {out_path}')
    print(f'   크기: {size_kb:.1f} KB / 시트: {len(sheets)}개')
    print('=' * 60)


if __name__ == '__main__':
    main()

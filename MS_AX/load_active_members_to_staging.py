import json
import os
import sys
import urllib.error
import urllib.request
import datetime

from openpyxl import load_workbook


XLSX_PATH = r"C:\csh\project\chflow\MS_AX\extracted-bathelDB\kyoin2015_member_segments.xlsx"
SHEET_NAME = "active_members"
ENV_PATH = r"C:\csh\project\chflow\chflow-app\.env.local"
TABLE = "staging_members_mdb"
CHUNK_SIZE = 300


def load_env(path):
    env = {}
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                env[k] = v
    return env


ENV = load_env(ENV_PATH)
SUPABASE_URL = ENV["NEXT_PUBLIC_SUPABASE_URL"]
SERVICE_KEY = ENV["SUPABASE_SERVICE_ROLE_KEY"]
HEADERS = {
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}


def http(method, path, body=None, headers=None):
    url = f"{SUPABASE_URL}/rest/v1{path}"
    data = json.dumps(body, ensure_ascii=False, default=str).encode("utf-8") if body is not None else None
    req = urllib.request.Request(url, data=data, headers={**HEADERS, **(headers or {})}, method=method)
    with urllib.request.urlopen(req, timeout=120) as resp:
        raw = resp.read()
        return json.loads(raw.decode("utf-8")) if raw else None


def chunked(rows, size):
    for i in range(0, len(rows), size):
        yield rows[i:i + size]


def norm_text(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def norm_bool(v):
    s = norm_text(v)
    if s is None:
        return None
    return s.lower() == "true"


def norm_date(v):
    s = norm_text(v)
    if not s or s == "-  -":
        return None
    try:
        return datetime.date.fromisoformat(s).isoformat()
    except Exception:
        return None
    return None


def read_active_rows():
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    rows = ws.iter_rows(values_only=True)
    headers = [str(v).strip() if v is not None else "" for v in next(rows)]
    out = []
    for idx, row in enumerate(rows, start=2):
        src = dict(zip(headers, row))
        out.append((idx, src))
    return out


def build_payload(source_row_no, src):
    row = {
        "source_system": "kyoin2015_mdb",
        "source_file": os.path.basename(XLSX_PATH),
        "source_row_no": source_row_no,
        "legacy_kyoin_id": norm_text(src.get("kyoinID")),
        "legacy_family_num": norm_text(src.get("family_num")),
        "legacy_seq_num": norm_text(src.get("seq_num")),
        "name": norm_text(src.get("name")),
        "english_name": norm_text(src.get("ename")),
        "birth_date": norm_date(src.get("birth")),
        "birth_raw": norm_text(src.get("birth")),
        "age_raw": norm_text(src.get("Age")),
        "gender": norm_text(src.get("mw")),
        "lunar_solar": norm_text(src.get("pm")),
        "relationship_in_household": norm_text(src.get("isa")),
        "phone": norm_text(src.get("tel")),
        "post_num": norm_text(src.get("post_num")),
        "address_line_1": norm_text(src.get("juso")),
        "address_line_2": norm_text(src.get("juso2")),
        "life_raw": norm_text(src.get("life")),
        "first_registration_date": norm_date(src.get("i_date")),
        "baptism_date": norm_date(src.get("s_date")),
        "other_date": norm_date(src.get("se_date")),
        "yang_raw": norm_text(src.get("Yang")),
        "yangi_raw": norm_text(src.get("Yangi")),
        "office_role_raw": norm_text(src.get("junPos")),
        "office_role_date_raw": norm_text(src.get("junDate")),
        "office_role_note_raw": norm_text(src.get("junFor")),
        "picture_ref": norm_text(src.get("picture1")),
        "sms_opt_in_raw": norm_text(src.get("sms")),
        "act_raw": norm_text(src.get("act")),
        "jong_raw": norm_text(src.get("jong")),
        "move_out": norm_bool(src.get("MoveOut")),
        "deleted_flag": norm_bool(src.get("del1")),
        "delete_reason": norm_text(src.get("delWhy")),
        "raw_payload": src,
    }
    return row


def main():
    source_rows = read_active_rows()
    payload = [build_payload(row_no, src) for row_no, src in source_rows]
    payload = [row for row in payload if row["name"]]

    print(f"active rows to load: {len(payload)}")
    http("DELETE", f"/{TABLE}?id=gt.0")
    print("staging cleared")

    inserted = 0
    for batch in chunked(payload, CHUNK_SIZE):
        http("POST", f"/{TABLE}", batch)
        inserted += len(batch)
        print(f"inserted {inserted}/{len(payload)}")

    print("done")


if __name__ == "__main__":
    try:
        main()
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        print(f"HTTP {e.code}: {body}")
        sys.exit(1)
    except Exception as e:
        print(str(e))
        sys.exit(1)

"""
사진 페이지 파서
패턴:
  이름 (배우자)
  자택전화
  휴대폰
  평원-초원-목장
"""
import sys, io, re, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import fitz
import openpyxl
from openpyxl.styles import Font, PatternFill

PDF = 'c:/csh/project/chflow/자료/데이터베이스 추출용(ocr을 텍스트버전으로변경).pdf'
OUT_DIR = 'C:/csh/project/chflow/MS_AX/parsed-data'
os.makedirs(OUT_DIR, exist_ok=True)


def normalize_phone(text):
    if not text:
        return ''
    cleaned = re.sub(r'\s+', '', text)
    m = re.search(r'(\d{2,3})-?(\d{3,4})-?(\d{4})', cleaned)
    return f'{m.group(1)}-{m.group(2)}-{m.group(3)}' if m else ''


def is_phone(text):
    return bool(re.search(r'\d{2,3}\s*-\s*\d{3,4}\s*-\s*\d{4}', text))


def parse_pasture_label(text):
    """3-신위식-장영국목장 → (평원=3, 초원=신위식, 목장=장영국)"""
    text = text.strip()
    text = re.sub(r'^\d+_', '', text)  # leading digit_
    # Try multi-format: "3-신위식-장영국목장" or "박두재_김정배목장"
    parts = re.split(r'[-_]', text)
    parts = [p.strip() for p in parts if p.strip()]
    if len(parts) == 3:
        return parts[0], parts[1], parts[2].replace('목장', '')
    if len(parts) == 2:
        return None, parts[0], parts[1].replace('목장', '')
    return None, None, text


def parse_name_spouse(text):
    """장영국(곽순이) → (장영국, 곽순이)"""
    text = text.strip()
    m = re.match(r'([가-힣]{2,4})\s*[\(\[]\s*([가-힣]{2,4})\s*[\)\]]', text)
    if m:
        return m.group(1), m.group(2)
    return text, None


def get_spans(page):
    spans = []
    d = page.get_text("dict")
    for block in d['blocks']:
        if block['type'] != 0:
            continue
        for line in block['lines']:
            for span in line['spans']:
                text = span['text'].strip()
                if text:
                    bbox = span['bbox']
                    spans.append({'text': text, 'x': bbox[0], 'y': bbox[1]})
    return sorted(spans, key=lambda s: (s['y'], s['x']))


def parse_photo_page(page, page_num):
    """사진 페이지에서 회원 정보 추출"""
    spans = get_spans(page)
    if not spans:
        return []

    # 각 회원 데이터: 이름(배우자) → 자택 → 휴대폰 → 평원-초원-목장
    # 4명이 한 가로 행으로 배치 (각 컬럼은 x 위치로 구분)
    # 이름 라인의 y를 anchor로 사용

    members = []
    used_y = set()

    # 이름 패턴 (한글 + 괄호 안 한글)
    name_anchors = [s for s in spans
                    if re.match(r'[가-힣]{2,4}\s*[\(\[].*[\)\]]', s['text'])
                    and s['y'] > 200]  # skip page header

    for anchor in name_anchors:
        y_anchor = anchor['y']
        x_anchor = anchor['x']

        # 같은 컬럼 내에서 이 이름 아래쪽 30px 이내의 span 찾기
        nearby = [s for s in spans
                  if abs(s['x'] - x_anchor) < 60
                  and 0 <= s['y'] - y_anchor <= 35]

        # 패턴: 이름 → 자택(7자리) → 휴대폰(11자리) → 평원-초원-목장
        name, spouse = parse_name_spouse(anchor['text'])
        home_phone = None
        mobile = None
        plain = grassland = pasture = None

        for s in nearby:
            t = s['text']
            if t == anchor['text']:
                continue
            if is_phone(t):
                # 010 으로 시작 → 휴대폰
                if t.lstrip().startswith('010'):
                    if not mobile:
                        mobile = normalize_phone(t)
                else:
                    if not home_phone:
                        home_phone = normalize_phone(t)
            elif re.search(r'목장', t):
                p, g, ps = parse_pasture_label(t)
                if p:
                    plain = p
                if g:
                    grassland = g
                if ps:
                    pasture = ps
            elif re.match(r'^\d{3}\s*-\s*\d{4}', t):
                # 자택 전화 (지역번호 없는 것)
                if not home_phone:
                    home_phone = normalize_phone(t)

        members.append({
            'name': name,
            'spouse': spouse,
            'home_phone': home_phone,
            'mobile': mobile,
            'plain': plain,
            'grassland': grassland,
            'pasture': pasture,
            'page': page_num,
        })

    return members


def main():
    doc = fitz.open(PDF)
    test_page = 10  # PDF page 11 (was showing photos)
    page = doc[test_page]
    print(f'=== Photo page {test_page + 1} ===')
    members = parse_photo_page(page, test_page + 1)
    print(f'\n{len(members)} members found:\n')

    for m in members:
        spouse_str = f"({m['spouse']})" if m['spouse'] else ''
        print(f"  {m['name']}{spouse_str}")
        print(f"    자택: {m['home_phone']} / 휴대: {m['mobile']}")
        print(f"    {m['plain']}평원 - {m['grassland']}초원 - {m['pasture']}목장")
        print()

    # Save to Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '사진페이지'
    ws.append(['이름', '배우자', '자택전화', '휴대폰', '평원', '초원', '목장', 'PDF페이지'])
    for c in range(1, 9):
        ws.cell(1, c).font = Font(bold=True)
        ws.cell(1, c).fill = PatternFill('solid', fgColor='E0E7FF')

    for m in members:
        ws.append([
            m['name'], m['spouse'] or '', m['home_phone'] or '',
            m['mobile'] or '', m['plain'] or '', m['grassland'] or '',
            m['pasture'] or '', m['page']
        ])

    widths = [12, 12, 14, 16, 8, 14, 14, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    out = os.path.join(OUT_DIR, 'test_page11_사진페이지.xlsx')
    wb.save(out)
    print(f'💾 Excel: {out}')


if __name__ == '__main__':
    main()

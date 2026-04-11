"""
명성교회 요람 PDF → 목장 데이터 파서 (V2)
주소 라인을 가족 경계로 사용하는 anchor-based 파서
"""
import sys, io, re, os, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import fitz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

PDF = 'c:/csh/project/chflow/자료/데이터베이스 추출용(ocr을 텍스트버전으로변경).pdf'
OUT_DIR = 'C:/csh/project/chflow/MS_AX/parsed-data'
os.makedirs(OUT_DIR, exist_ok=True)


# ============================================================
# 알려진 직분 (긴 것부터 매칭)
# ============================================================
ROLES_KNOWN = sorted([
    '담임목사', '부목사', '은퇴목사', '목사',
    '선교사', '전도사', '사모',
    '시무장로', '원로장로', '은퇴장로', '명예장로', '장로',
    '교육사', '간사',
    '명예시무집사', '은퇴시무집사', '시무집사',
    '명예시무권사', '은퇴시무권사', '시무권사',
    '명예집사', '은퇴집사', '집사',
    '명예권사', '은퇴권사', '권사',
    '청년', '청소년', '어린이', '유아',
], key=len, reverse=True)

# 지역 키워드 (주소 식별용)
ADDRESS_KEYWORDS = ['동구', '북구', '남구', '중구', '울주군', '경북', '서울', '경기',
                    '부산', '대구', '광주', '대전', '인천', '울산']

OCR_FIXES = {
    '권Af': '권사', '권f': '권사',
    '시T집사': '시무집사', '지T집사': '시무집사',
    'X[': '자', '0 1': '01',
}


def fix_ocr(text):
    if not text:
        return ''
    for bad, good in OCR_FIXES.items():
        text = text.replace(bad, good)
    return text


def normalize_phone(text):
    if not text:
        return ''
    cleaned = re.sub(r'\s+', '', text)
    m = re.search(r'(\d{2,3})-?(\d{3,4})-?(\d{4})', cleaned)
    return f'{m.group(1)}-{m.group(2)}-{m.group(3)}' if m else cleaned


def extract_role(text):
    text = fix_ocr(text)
    for role in ROLES_KNOWN:
        if role in text:
            return role
    return None


def is_address_line(text):
    text = text.strip()
    return any(text.startswith(k) or k in text[:6] for k in ADDRESS_KEYWORDS)


def is_phone_text(text):
    return bool(re.search(r'\d{2,3}\s*[-]\s*\d{3,4}\s*[-]\s*\d{4}', text))


def get_words(page):
    words = []
    d = page.get_text("dict")
    for block in d['blocks']:
        if block['type'] != 0:
            continue
        for line in block['lines']:
            for span in line['spans']:
                text = span['text'].strip()
                if text:
                    bbox = span['bbox']
                    words.append({
                        'text': fix_ocr(text),
                        'x': bbox[0],
                        'y': bbox[1],
                        'x2': bbox[2],
                        'y2': bbox[3],
                    })
    return sorted(words, key=lambda w: (w['y'], w['x']))


def parse_household_block(block_words, no, addr_col_x):
    """한 가족 영역의 단어들에서 정보 추출"""
    names = []
    roles = []
    home_phone = None
    mobile_phones = []
    address_parts = []
    children_parts = []

    # 컬럼별 분류 (헤더 위치 기반 추정)
    name_col_max = addr_col_x - 30
    role_col_max = addr_col_x - 5
    phone_col_min = addr_col_x + 110

    for w in block_words:
        x, text = w['x'], w['text']

        # 전화 컬럼
        if x >= phone_col_min:
            if 'T.' in text or 'T ' in text:
                home_phone = normalize_phone(text)
            elif is_phone_text(text):
                # 한국어 첫글자 + ":" + 전화 → 개인 휴대폰
                m = re.search(r'(\d{2,3}[-\s]?\d{3,4}[-\s]?\d{4})', text)
                if m:
                    mobile_phones.append(normalize_phone(m.group(1)))
            continue

        # 주소/가족 컬럼
        if x >= addr_col_x - 5 and x < phone_col_min:
            # 텍스트 분리 시도 (한 셀에 주소+자녀가 있을 수 있음)
            t = text.strip()
            if not t:
                continue
            if is_address_line(t):
                address_parts.append(t)
            elif any(k in t[:6] for k in ['(', '아파트', '빌라', '동', '호']):
                # 주소의 연속 (서부동, 서부아파트) 같은 것
                address_parts.append(t)
            else:
                # 가족(자녀)
                # 이름 패턴: 한글 2~4자
                children_parts.append(t)
            continue

        # 직분 컬럼
        if name_col_max <= x < role_col_max:
            t = text.strip()
            # "심71욱 |은퇴장로|" 같은 합쳐진 텍스트 처리
            for part in re.split(r'[\|]', t):
                part = part.strip()
                if not part:
                    continue
                role = extract_role(part)
                if role:
                    if role not in roles:
                        roles.append(role)
                else:
                    # 이름일 수도 있음
                    if re.match(r'^[가-힣]{2,4}$', part) and part not in names:
                        names.append(part)
            continue

        # 성명 컬럼
        if x < name_col_max:
            t = text.strip()
            # "1 | 김OO" 같은 형태
            t = re.sub(r'^\d+\s*[\|\.]?\s*', '', t)
            for part in re.split(r'[\|]', t):
                part = part.strip()
                if not part:
                    continue
                # 직분이 섞여있을 수 있음
                role = extract_role(part)
                if role:
                    if role not in roles:
                        roles.append(role)
                    # 직분 뒤에 이름이 더 있을 수도
                    rest = part.replace(role, '').strip()
                    if re.match(r'^[가-힣]{2,4}$', rest):
                        if rest not in names:
                            names.append(rest)
                elif re.match(r'^[가-힣]{2,4}$', part):
                    if part not in names:
                        names.append(part)

    return {
        'no': no,
        'names': names[:2],  # 부부 최대 2명
        'roles': roles[:2],
        'address': ' '.join(address_parts).strip(),
        'children': children_parts,
        'phone_home': home_phone,
        'phone_mobiles': mobile_phones,
    }


def parse_pasture_page(page, page_num):
    words = get_words(page)
    if not words:
        return None

    # 헤더와 본문 분리 (목자/목녀 영역은 y < 230)
    header_words = [w for w in words if w['y'] < 230]
    body_words = [w for w in words if w['y'] >= 230]

    # 컬럼 위치 추정 (헤더 행에서)
    addr_col_x = 248  # 기본값
    for w in body_words:
        if w['text'].strip() == '주소':
            addr_col_x = max(addr_col_x, int(w['x']) - 5)
            break
    # 실제 데이터 위치로 보정 (헤더 위치보다 조금 왼쪽)
    addr_col_x = 245

    # 헤더(목자/목녀) 파싱
    shepherd = shepherdess = shepherd_phone = shepherdess_phone = None
    for w in header_words:
        t = w['text']
        if w['x'] < 350:  # 좌측 = 목자
            if re.match(r'^[가-힣]{2,4}$', t) and shepherd is None:
                shepherd = t
            elif is_phone_text(t):
                shepherd_phone = normalize_phone(t)
        else:  # 우측 = 목녀
            if re.match(r'^[가-힣]{2,4}$', t) and shepherdess is None and t != '목녀':
                shepherdess = t
            elif is_phone_text(t):
                shepherdess_phone = normalize_phone(t)

    # 가족 경계 찾기: 주소 라인의 y 좌표
    address_anchors = []
    for w in body_words:
        x, t = w['x'], w['text']
        if x >= addr_col_x - 10 and x < addr_col_x + 130:
            if is_address_line(t):
                address_anchors.append({'y': w['y'], 'text': t})

    # y로 정렬, 중복 제거 (가까운 y는 같은 주소의 연속)
    address_anchors.sort(key=lambda a: a['y'])
    unique_anchors = []
    for a in address_anchors:
        if not unique_anchors or a['y'] - unique_anchors[-1]['y'] > 20:
            unique_anchors.append(a)

    # 각 가족 영역 파싱
    households = []
    for i, anchor in enumerate(unique_anchors):
        y_start = anchor['y'] - 10
        y_end = unique_anchors[i + 1]['y'] - 10 if i + 1 < len(unique_anchors) else 1000
        block_words = [w for w in body_words if y_start <= w['y'] < y_end]
        h = parse_household_block(block_words, i + 1, addr_col_x)
        households.append(h)

    return {
        'page': page_num,
        'shepherd': shepherd,
        'shepherdess': shepherdess,
        'shepherd_phone': shepherd_phone,
        'shepherdess_phone': shepherdess_phone,
        'households': households,
    }


def main():
    doc = fitz.open(PDF)
    print(f'Total pages: {doc.page_count}')

    test_page = 50  # PDF p.51
    page = doc[test_page]
    print(f'\n=== Parsing PDF page {test_page + 1} ===')
    result = parse_pasture_page(page, test_page + 1)

    print(f"\n목자: {result['shepherd']} ({result['shepherd_phone']})")
    print(f"목녀: {result['shepherdess']} ({result['shepherdess_phone']})")
    print(f"가족 수: {len(result['households'])}")
    print()

    for h in result['households']:
        print(f"  [{h['no']}] {' / '.join(h['names'])}")
        print(f"      직분: {' / '.join(h['roles'])}")
        print(f"      주소: {h['address']}")
        if h['children']:
            print(f"      자녀: {', '.join(h['children'])}")
        print(f"      자택: {h['phone_home']}")
        if h['phone_mobiles']:
            print(f"      휴대: {', '.join(h['phone_mobiles'])}")
        print()

    # Excel 출력
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '심기욱목장'

    ws.append(['평원', '초원', '목장', '목자/목녀', '이름', '직분', '주소', '자녀', '자택', '휴대폰'])
    for col in range(1, 11):
        ws.cell(1, col).font = Font(bold=True)
        ws.cell(1, col).fill = PatternFill('solid', fgColor='E0E7FF')

    # 목자/목녀 행 (수동 입력 필요한 평원/초원/목장)
    pasture_label = f"{result['shepherd']}목장" if result['shepherd'] else ''
    ws.append([
        '1평원', '캄보디아', pasture_label,
        '목자', result['shepherd'] or '', '', '', '', '', result['shepherd_phone'] or ''
    ])
    if result['shepherdess']:
        ws.append([
            '', '', '',
            '목녀', result['shepherdess'], '', '', '', '', result['shepherdess_phone'] or ''
        ])

    # 본문
    for h in result['households']:
        names = h['names'] or ['']
        roles = h['roles'] or ['']
        children = ', '.join(h['children']) if h['children'] else ''
        for i, name in enumerate(names):
            role = roles[i] if i < len(roles) else ''
            mob = h['phone_mobiles'][i] if i < len(h['phone_mobiles']) else ''
            ws.append([
                '', '', '',
                f'#{h["no"]}' if i == 0 else '',
                name,
                role,
                h['address'] if i == 0 else '',
                children if i == 0 else '',
                h['phone_home'] or '' if i == 0 else '',
                mob,
            ])

    widths = [8, 12, 14, 10, 12, 14, 45, 25, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    out = os.path.join(OUT_DIR, 'test_page51_심기욱목장.xlsx')
    wb.save(out)
    print(f'\n💾 Excel: {out}')


if __name__ == '__main__':
    main()

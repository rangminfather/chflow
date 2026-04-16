"""
fallback 매칭:
- 각 PDF 사진페이지에서 이미지들의 bbox를 reading order(y→x)로 정렬
- 같은 페이지의 엑셀 row 순서와 1:1 대응 (개수 일치할 때만)
- 이미 photo_url 있는 회원은 skip
"""
import sys, io, os, re, json, urllib.request, urllib.error, urllib.parse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from collections import defaultdict
from openpyxl import load_workbook
import fitz

PDF        = 'c:/csh/project/chflow/자료/데이터베이스 추출용(ocr을 텍스트버전으로변경).pdf'
EXCEL      = 'c:/csh/project/chflow/test-excel/test_2_fixed_v2.xlsx'
PHOTOS_DIR = 'c:/csh/project/chflow/MS_AX/parsed-data/photos'
ENV        = 'c:/csh/project/chflow/chflow-app/.env.local'
BUCKET     = 'member-photos'

with open(ENV, 'r', encoding='utf-8') as f:
    for line in f:
        line = line.strip()
        if '=' in line and not line.startswith('#'):
            k, v = line.split('=', 1); os.environ[k] = v
URL = os.environ['NEXT_PUBLIC_SUPABASE_URL']
KEY = os.environ['SUPABASE_SERVICE_ROLE_KEY']
H   = {'apikey': KEY, 'Authorization': f'Bearer {KEY}'}


def _enc(p):
    if '?' in p:
        b, q = p.split('?', 1); return b + '?' + urllib.parse.quote(q, safe='=&.,%*!')
    return p

def get_all(path):
    out, off, step = [], 0, 1000
    while True:
        sep = '&' if '?' in path else '?'
        req = urllib.request.Request(f'{URL}/rest/v1{_enc(path+sep+f"offset={off}&limit={step}")}', headers=H)
        data = json.loads(urllib.request.urlopen(req, timeout=60).read())
        out.extend(data)
        if len(data) < step: break
        off += step
    return out

def norm_phone(s):
    return re.sub(r'\D', '', s or '')


# ── 엑셀: 페이지별 row 순서 ──
wb = load_workbook(EXCEL); ws = wb.active
by_page = defaultdict(list)  # page → [(excel_idx, name, phone)]
for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True)):
    if r[18] is None: continue
    by_page[int(r[18])].append((i, (r[6] or '').strip(), (r[16] or '').strip()))
for p in by_page:
    by_page[p].sort(key=lambda t: t[0])

# ── DB members ──
members = get_all('/members?select=id,name,phone,is_child,photo_url')
by_name = defaultdict(list)
for m in members: by_name[m['name']].append(m)

def find_member(name, phone):
    pool = [m for m in by_name.get(name, []) if not m['is_child']]
    if not pool: pool = by_name.get(name, [])
    if not pool: return None
    if phone:
        exact = [m for m in pool if norm_phone(m['phone']) == norm_phone(phone)]
        if exact: return exact[0]
        last4 = [m for m in pool if norm_phone(m['phone'])[-4:] == norm_phone(phone)[-4:]]
        if len(last4) == 1: return last4[0]
    if len(pool) == 1: return pool[0]
    return None

# ── PDF: page별 photo reading-order 정렬 ──
doc = fitz.open(PDF)
ready = []          # [(member_id, file_path, name)]
stats = {'skip_count_mismatch':0, 'already_has':0, 'not_found':0, 'queued':0}

for page_idx in range(43):
    page = doc[page_idx]
    img_bboxes = {inf['xref']: inf['bbox'] for inf in page.get_image_info(xrefs=True)}
    photos = []  # extract_all.py와 동일한 photo_idx 순서
    pcount = 0
    for img in page.get_images(full=True):
        xref, _sm, w, h = img[0], img[1], img[2], img[3]
        if not (400 < w < 700 and 400 < h < 700): continue
        bb = img_bboxes.get(xref)
        if not bb: continue
        photos.append({'photo_idx': pcount, 'bbox': bb})
        pcount += 1

    # y 버킷(60pt) → x 로 reading order 정렬
    photos_ro = sorted(photos, key=lambda p: (round(p['bbox'][1] / 60.0), p['bbox'][0]))

    excel_list = by_page.get(page_idx + 1, [])
    if len(photos_ro) != len(excel_list):
        if excel_list:
            stats['skip_count_mismatch'] += len(excel_list)
            print(f'  p{page_idx+1:03d}: photos={len(photos_ro)} vs excel={len(excel_list)} → skip')
        continue

    for ph, (eidx, name, phone) in zip(photos_ro, excel_list):
        fname = f'p{page_idx+1:03d}_photo{ph["photo_idx"]:02d}.png'
        fpath = os.path.join(PHOTOS_DIR, fname)
        if not os.path.exists(fpath): continue
        m = find_member(name, phone)
        if not m:
            stats['not_found'] += 1; continue
        if m.get('photo_url'):
            stats['already_has'] += 1; continue
        ready.append((m['id'], fpath, name))
        stats['queued'] += 1

print(f'\n통계: {stats}')
print(f'추가 업로드 대상: {len(ready)}')


# ── 업로드 + UPDATE ──
ok, fail = 0, 0
for i, (mid, fpath, name) in enumerate(ready, 1):
    with open(fpath, 'rb') as f: data = f.read()
    obj = f'{mid}/profile.png'
    up_url = f'{URL}/storage/v1/object/{BUCKET}/{obj}'
    req = urllib.request.Request(up_url, data=data, method='POST',
        headers={**H, 'Content-Type': 'image/png', 'x-upsert': 'true'})
    try:
        urllib.request.urlopen(req, timeout=30).read()
    except urllib.error.HTTPError:
        try:
            req2 = urllib.request.Request(up_url, data=data, method='PUT',
                headers={**H, 'Content-Type': 'image/png', 'x-upsert': 'true'})
            urllib.request.urlopen(req2, timeout=30).read()
        except Exception:
            fail += 1; continue

    patch_url  = f'{URL}/rest/v1/members?id=eq.{mid}'
    public_url = f'{URL}/storage/v1/object/public/{BUCKET}/{obj}'
    req = urllib.request.Request(patch_url,
        data=json.dumps({'photo_url': public_url}).encode('utf-8'),
        method='PATCH',
        headers={**H, 'Content-Type': 'application/json', 'Prefer': 'return=minimal'})
    try:
        urllib.request.urlopen(req, timeout=30).read(); ok += 1
    except Exception:
        fail += 1
    if i % 25 == 0:
        print(f'  진행 {i}/{len(ready)} (성공 {ok}, 실패 {fail})')

print(f'\n성공 {ok}, 실패 {fail}')
